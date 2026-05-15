"""
Airline Accounting Real-Time Excel Generator
============================================
Standards applied: IAS 21 / IFRS 21, IFRS 16, Ind AS 116, Ind AS 11, IATA AAG
Base currency: EUR  |  Parallel ledgers: USD, GBP, INR, SGD

Usage:
    python airline_accounting.py [--interval N] [--output FILE] [--no-api]

Press Ctrl+C to stop.
"""

import argparse
import calendar
import dataclasses
import datetime
import json
import os
import random
import sys
import time
from decimal import Decimal, getcontext, ROUND_HALF_UP
from typing import Dict, List, Optional, Tuple

getcontext().prec = 12

try:
    import requests
    _REQUESTS_AVAILABLE = True
except ImportError:
    _REQUESTS_AVAILABLE = False

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
BASE_CURRENCY       = "EUR"
FOREIGN_CURRENCIES  = ["USD", "GBP", "INR", "SGD"]
REFRESH_INTERVAL    = 15
OUTPUT_FILE         = "airline_accounting.xlsx"
FX_API_URL          = "https://api.frankfurter.dev/v2/rates"
FX_API_TIMEOUT      = 3

_today = datetime.date.today()
FISCAL_YEAR_START   = datetime.date(_today.year, 1, 1)
DAYS_IN_YEAR        = 366 if calendar.isleap(_today.year) else 365

# Colours
AIRLINE_BLUE   = "003087"
HEADER_GREEN   = "1F5C1F"
SECTION_PURPLE = "4B0082"
ALT_ROW_1      = "DCE6F1"
ALT_ROW_2      = "FFFFFF"
RED_FILL_HEX   = "FFCCCC"
GREEN_FILL_HEX = "CCFFCC"
AMBER_FILL_HEX = "FFF2CC"
GOLD_FILL_HEX  = "FFD700"
TEAL_FILL_HEX  = "008080"

# Base FX rates (EUR base) for simulation seed
_BASE_RATES: Dict[str, Decimal] = {
    "USD": Decimal("1.0823"),
    "GBP": Decimal("0.8561"),
    "INR": Decimal("90.12"),
    "SGD": Decimal("1.4672"),
}

SHEET_NAMES = [
    "Dashboard",
    "General Ledger EUR",
    "Parallel Ledger USD",
    "Parallel Ledger GBP",
    "Parallel Ledger INR",
    "Parallel Ledger SGD",
    "Forex Revaluation",
    "Interest Calculation",
    "IFRS 16 Lease Schedule",
    "Air India A320 Lease",
    "Year-End Closing",
    "Standards Reference",
    "Airline Lease Comparison",
]

_TWO = Decimal("0.01")
_FOUR = Decimal("0.0001")


def _d2(v: Decimal) -> Decimal:
    return v.quantize(_TWO, rounding=ROUND_HALF_UP)


def _d4(v: Decimal) -> Decimal:
    return v.quantize(_FOUR, rounding=ROUND_HALF_UP)


# ---------------------------------------------------------------------------
# Transaction dataclass
# ---------------------------------------------------------------------------
@dataclasses.dataclass
class Transaction:
    tx_id:        str
    date:         datetime.date
    description:  str
    account:      str
    account_type: str          # ASSET | LIABILITY | EQUITY | REVENUE | EXPENSE
    debit_eur:    Decimal
    credit_eur:   Decimal
    currency:     str
    fx_rate:      Decimal      # EUR → FCY rate at transaction date
    is_monetary:  bool         = True
    lease_id:     Optional[str] = None
    debit_fcy:    Decimal      = dataclasses.field(init=False)
    credit_fcy:   Decimal      = dataclasses.field(init=False)

    def __post_init__(self):
        if self.debit_eur < 0 or self.credit_eur < 0:
            raise ValueError(f"Negative amount in {self.tx_id}")
        if self.debit_eur == 0 and self.credit_eur == 0:
            raise ValueError(f"Both debit and credit are zero in {self.tx_id}")
        self.debit_fcy  = _d2(self.debit_eur  * self.fx_rate)
        self.credit_fcy = _d2(self.credit_eur * self.fx_rate)

    @property
    def net_eur(self) -> Decimal:
        return self.debit_eur - self.credit_eur


# ---------------------------------------------------------------------------
# FX Rate Manager
# ---------------------------------------------------------------------------
class FXRateManager:
    def __init__(self, use_live: bool = True):
        self._current: Dict[str, Decimal] = dict(_BASE_RATES)
        self._history: Dict[str, List[Decimal]] = {c: [] for c in FOREIGN_CURRENCIES}
        self._use_live = use_live and _REQUESTS_AVAILABLE
        self._year_opening: Optional[Dict[str, Decimal]] = None

    def fetch_or_simulate(self) -> Dict[str, Decimal]:
        if self._use_live:
            try:
                rates = self._fetch_live()
                self._current = rates
                return rates
            except Exception:
                self._use_live = False
        rates = self._simulate()
        self._current = rates
        return rates

    def _fetch_live(self) -> Dict[str, Decimal]:
        symbols = ",".join(FOREIGN_CURRENCIES)
        url = f"{FX_API_URL}?base={BASE_CURRENCY}&quotes={symbols}"
        resp = requests.get(url, timeout=FX_API_TIMEOUT)
        resp.raise_for_status()
        data = resp.json()
        result: Dict[str, Decimal] = {}
        for ccy in FOREIGN_CURRENCIES:
            result[ccy] = Decimal(str(data["rates"][ccy]))
        return result

    def _simulate(self) -> Dict[str, Decimal]:
        result: Dict[str, Decimal] = {}
        for ccy, rate in self._current.items():
            delta = Decimal(str(random.gauss(0, 0.0008)))
            new_rate = rate * (Decimal("1") + delta)
            base = _BASE_RATES[ccy]
            lo = base * Decimal("0.97")
            hi = base * Decimal("1.03")
            new_rate = max(lo, min(hi, new_rate))
            result[ccy] = _d4(new_rate)
        return result

    def record_rates(self, rates: Dict[str, Decimal]) -> None:
        for ccy, rate in rates.items():
            if ccy in self._history:
                self._history[ccy].append(rate)
        if self._year_opening is None:
            self._year_opening = dict(rates)

    def average_rates(self) -> Dict[str, Decimal]:
        result: Dict[str, Decimal] = {}
        for ccy in FOREIGN_CURRENCIES:
            hist = self._history[ccy]
            if hist:
                result[ccy] = _d4(Decimal(str(sum(hist) / len(hist))))
            else:
                result[ccy] = self._current[ccy]
        return result

    def closing_rates(self) -> Dict[str, Decimal]:
        return dict(self._current)

    def year_opening_rates(self) -> Dict[str, Decimal]:
        if self._year_opening:
            return dict(self._year_opening)
        return dict(_BASE_RATES)

    def is_live(self) -> bool:
        return self._use_live


# ---------------------------------------------------------------------------
# Ledger Book
# ---------------------------------------------------------------------------
class LedgerBook:
    def __init__(self):
        self._transactions: List[Transaction] = []
        self._counter = 0

    def next_tx_id(self) -> str:
        self._counter += 1
        return f"TX-{self._counter:04d}"

    def add_transaction(self, tx: Transaction) -> None:
        self._transactions.append(tx)

    def all_transactions(self) -> List[Transaction]:
        return sorted(self._transactions, key=lambda t: (t.date, t.tx_id))

    def balance_by_account(self) -> Dict[str, Decimal]:
        balances: Dict[str, Decimal] = {}
        for tx in self._transactions:
            balances[tx.account] = balances.get(tx.account, Decimal("0")) + tx.net_eur
        return balances

    def monetary_balances_fcy(self) -> Dict[str, Dict[str, Decimal]]:
        result: Dict[str, Dict[str, Decimal]] = {}
        for tx in self._transactions:
            if not tx.is_monetary or tx.currency not in FOREIGN_CURRENCIES:
                continue
            ccy_map = result.setdefault(tx.currency, {})
            ccy_map[tx.account] = ccy_map.get(tx.account, Decimal("0")) + (tx.debit_fcy - tx.credit_fcy)
        return result

    def total_assets(self) -> Decimal:
        return sum(
            b for acct, b in self.balance_by_account().items()
            if _account_type(acct) == "ASSET"
        ) or Decimal("0")

    def total_liabilities(self) -> Decimal:
        return abs(sum(
            b for acct, b in self.balance_by_account().items()
            if _account_type(acct) == "LIABILITY"
        )) or Decimal("0")

    def total_revenue(self) -> Decimal:
        return abs(sum(
            b for acct, b in self.balance_by_account().items()
            if _account_type(acct) == "REVENUE"
        )) or Decimal("0")

    def total_expenses(self) -> Decimal:
        return sum(
            b for acct, b in self.balance_by_account().items()
            if _account_type(acct) == "EXPENSE"
        ) or Decimal("0")

    def net_income(self) -> Decimal:
        return self.total_revenue() - self.total_expenses()


_ACCOUNT_TYPE_MAP: Dict[str, str] = {
    "Cash/Bank":                "ASSET",
    "Accounts Receivable":      "ASSET",
    "ROU Asset":                "ASSET",
    "ROU Asset Depreciation":   "EXPENSE",
    "Lease Liability":          "LIABILITY",
    "Maintenance Payable":      "LIABILITY",
    "Salaries Payable":         "LIABILITY",
    "ATC Charges Payable":      "LIABILITY",
    "Airport Charges Payable":  "LIABILITY",
    "Fuel Payable":             "LIABILITY",
    "Fuel Expense":             "EXPENSE",
    "Interest Expense":         "EXPENSE",
    "Maintenance Reserve":      "EXPENSE",
    "Staff Costs":              "EXPENSE",
    "Airport Charges":          "EXPENSE",
    "Ground Handling":          "EXPENSE",
    "Insurance Expense":        "EXPENSE",
    "Catering Costs":           "EXPENSE",
    "Navigation Fees":          "EXPENSE",
    "Passenger Revenue":        "REVENUE",
    "Cargo Revenue":            "REVENUE",
    "Fuel Hedge Gain":          "REVENUE",
    "Charter Revenue":          "REVENUE",
    "Ancillary Revenue":        "REVENUE",
    "FX Gain":                  "REVENUE",
    "FX Loss":                  "EXPENSE",
}


def _account_type(account: str) -> str:
    return _ACCOUNT_TYPE_MAP.get(account, "ASSET")


# ---------------------------------------------------------------------------
# IFRS 16 Lease Schedule
# ---------------------------------------------------------------------------
@dataclasses.dataclass
class LeaseConfig:
    lease_id:           str
    aircraft_type:      str
    commencement:       datetime.date
    lease_term_months:  int
    monthly_payment:    Decimal
    effective_rate:     Decimal   # annual, e.g. Decimal("0.065")
    residual_value:     Decimal   = Decimal("0")


class IFRS16LeaseSchedule:
    def __init__(self, config: LeaseConfig):
        self.config = config
        self._schedule: Optional[List[dict]] = None

    def monthly_rate(self) -> Decimal:
        r = (Decimal("1") + self.config.effective_rate) ** (Decimal("1") / Decimal("12")) - Decimal("1")
        return _d4(r)

    def initial_pv(self) -> Decimal:
        r = self.monthly_rate()
        n = self.config.lease_term_months
        p = self.config.monthly_payment
        if r == 0:
            pv = p * Decimal(str(n))
        else:
            pv = p * (Decimal("1") - (Decimal("1") + r) ** (-n)) / r
        return _d2(pv)

    def compute_schedule(self) -> List[dict]:
        if self._schedule is not None:
            return self._schedule
        schedule = []
        r = self.monthly_rate()
        pv = self.initial_pv()
        n = self.config.lease_term_months
        dep = _d2(pv / Decimal(str(n)))
        opening_liability = pv
        rou_opening = pv
        cum_interest = Decimal("0")
        cum_dep = Decimal("0")

        for i in range(1, n + 1):
            interest = _d2(opening_liability * r)
            principal = _d2(self.config.monthly_payment - interest)
            closing_liability = _d2(max(Decimal("0"), opening_liability - principal))
            rou_closing = _d2(max(Decimal("0"), rou_opening - dep))
            cum_interest += interest
            cum_dep += dep
            pay_date = self.config.commencement + datetime.timedelta(days=30 * i)
            schedule.append({
                "period":             i,
                "date":               pay_date,
                "opening_liability":  opening_liability,
                "interest_charge":    interest,
                "payment":            self.config.monthly_payment,
                "principal":          principal,
                "closing_liability":  closing_liability,
                "rou_opening":        rou_opening,
                "rou_depreciation":   dep,
                "rou_closing":        rou_closing,
                "cum_interest":       cum_interest,
                "cum_depreciation":   cum_dep,
            })
            opening_liability = closing_liability
            rou_opening = rou_closing

        self._schedule = schedule
        return schedule

    def current_period_idx(self) -> int:
        today = datetime.date.today()
        schedule = self.compute_schedule()
        for i, row in enumerate(schedule):
            if row["date"] > today:
                return i
        return len(schedule) - 1

    def current_liability(self) -> Decimal:
        idx = self.current_period_idx()
        return self.compute_schedule()[idx]["closing_liability"]

    def current_rou(self) -> Decimal:
        idx = self.current_period_idx()
        return self.compute_schedule()[idx]["rou_closing"]

    def cumulative_interest_ytd(self) -> Decimal:
        today = datetime.date.today()
        total = Decimal("0")
        for row in self.compute_schedule():
            if row["date"].year == today.year and row["date"] <= today:
                total += row["interest_charge"]
        return total


# ---------------------------------------------------------------------------
# Air India A320 lease parameters (real-world approximation)
# Source basis:
#   - Air India fleet: ~30 A320/A320neo family on operating lease (post-Tata acquisition)
#   - Typical narrow-body lease rate: USD 350,000–400,000/month (AVAC/IBA 2023-24 data)
#   - Lease term: 12 years (144 months) standard for new A320neo
#   - Incremental borrowing rate used as discount rate per Ind AS 116 para 26:
#     Air India credit rating ~BB, Indian sovereign + spread → ~8.5% p.a.
#   - USD/INR closing rate applied; EUR used as workbook base (USD lease denominated)
#   - Commencement: 01-Apr-2023 (post-Tata takeover, new fleet induction cycle)
#   - All figures in USD; converted to EUR in workbook at prevailing rate
#   - Maintenance reserves: ~USD 50,000/month (not included in lease liability per IFRS 16)
# ---------------------------------------------------------------------------
AI_A320_LEASE = LeaseConfig(
    lease_id          = "AI-A320NEO-2023",
    aircraft_type     = "Airbus A320neo (Air India - VT-RTA series)",
    commencement      = datetime.date(2023, 4, 1),
    lease_term_months = 144,              # 12 years
    monthly_payment   = Decimal("375000"),# USD 375,000/month (mid-range market rate)
    effective_rate    = Decimal("0.085"), # 8.5% p.a. — Ind AS 116 incremental borrowing rate
    residual_value    = Decimal("0"),     # Operating lease; no guaranteed residual
)


# ---------------------------------------------------------------------------
# Workbook Writer
# ---------------------------------------------------------------------------
class WorkbookWriter:
    def __init__(self, ledger: LedgerBook, fx: FXRateManager, leases: List[IFRS16LeaseSchedule]):
        self.ledger = ledger
        self.fx = fx
        self.leases = leases
        self._init_styles()

    # -- style helpers --
    def _init_styles(self):
        def fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        self.fill_blue    = fill(AIRLINE_BLUE)
        self.fill_green   = fill(HEADER_GREEN)
        self.fill_purple  = fill(SECTION_PURPLE)
        self.fill_alt1    = fill(ALT_ROW_1)
        self.fill_alt2    = fill(ALT_ROW_2)
        self.fill_red     = fill(RED_FILL_HEX)
        self.fill_green_l = fill(GREEN_FILL_HEX)
        self.fill_amber   = fill(AMBER_FILL_HEX)
        self.fill_gold    = fill(GOLD_FILL_HEX)
        self.fill_teal    = fill(TEAL_FILL_HEX)

        self.font_white_bold = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        self.font_hdr        = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        self.font_bold       = Font(name="Calibri", bold=True, size=10)
        self.font_normal     = Font(name="Calibri", size=10)
        self.font_small      = Font(name="Calibri", size=9, italic=True)

        thin = Side(border_style="thin", color="BBBBBB")
        self.border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
        thick_b = Side(border_style="medium", color="888888")
        self.border_bottom = Border(bottom=thick_b)

        self.align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
        self.align_right  = Alignment(horizontal="right",  vertical="center")
        self.align_left   = Alignment(horizontal="left",   vertical="center")
        self.align_wrap   = Alignment(horizontal="left",   vertical="top", wrap_text=True)

    def _hdr(self, cell, text: str = "", fill=None):
        cell.value = text
        cell.fill  = fill or self.fill_blue
        cell.font  = self.font_hdr
        cell.alignment = self.align_center
        cell.border = self.border_thin

    def _data(self, cell, value, fmt: str = None, bold: bool = False):
        cell.value = value
        cell.font  = self.font_bold if bold else self.font_normal
        cell.alignment = self.align_right if isinstance(value, (int, float, Decimal)) else self.align_left
        cell.border = self.border_thin
        if fmt:
            cell.number_format = fmt

    def _title_row(self, ws, text: str, last_col: int, fill=None):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        c = ws.cell(row=1, column=1)
        c.value = text
        c.fill  = fill or self.fill_blue
        c.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
        c.alignment = self.align_center
        ws.row_dimensions[1].height = 24

    def _last_updated(self, ws):
        r = ws.max_row + 1
        ws.cell(row=r, column=1).value = "Last Updated:"
        ws.cell(row=r, column=1).font = self.font_small
        ws.cell(row=r, column=2).value = datetime.datetime.now().strftime("%Y-%m-%d  %H:%M:%S")
        ws.cell(row=r, column=2).font = self.font_small

    def _alt_row(self, ws, row: int, max_col: int):
        fill = self.fill_alt1 if row % 2 == 0 else self.fill_alt2
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).fill = fill

    def _auto_width(self, ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val = str(cell.value) if cell.value is not None else ""
                    max_len = max(max_len, len(val))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)

    # -- main write entry point --
    def write(self, filepath: str) -> None:
        tmp = filepath + ".tmp"
        if os.path.exists(filepath):
            try:
                wb = openpyxl.load_workbook(filepath)
            except Exception:
                wb = openpyxl.Workbook()
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]
        else:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        for name in SHEET_NAMES:
            if name in wb.sheetnames:
                ws = wb[name]
                ws.delete_rows(1, ws.max_row + 1)
                for cf in list(ws.conditional_formatting._cf_rules.keys()):
                    del ws.conditional_formatting._cf_rules[cf]
            else:
                wb.create_sheet(name)

        # Reorder sheets
        for i, name in enumerate(SHEET_NAMES):
            if name in wb.sheetnames:
                wb.move_sheet(name, offset=wb.sheetnames.index(name) - i if name in wb.sheetnames else 0)

        self._write_dashboard(wb["Dashboard"])
        self._write_general_ledger(wb["General Ledger EUR"])
        for ccy in FOREIGN_CURRENCIES:
            self._write_parallel_ledger(wb[f"Parallel Ledger {ccy}"], ccy)
        self._write_forex_revaluation(wb["Forex Revaluation"])
        self._write_interest_calculation(wb["Interest Calculation"])
        self._write_ifrs16_schedule(wb["IFRS 16 Lease Schedule"])
        self._write_airindia_a320_lease(wb["Air India A320 Lease"])
        self._write_year_end_closing(wb["Year-End Closing"])
        self._write_standards_reference(wb["Standards Reference"])
        self._write_airline_lease_comparison(wb["Airline Lease Comparison"])

        wb.save(tmp)
        os.replace(tmp, filepath)

    # -----------------------------------------------------------------------
    # Sheet 1 — Dashboard
    # -----------------------------------------------------------------------
    def _write_dashboard(self, ws):
        self._title_row(ws, "AIRLINE ACCOUNTING SYSTEM — REAL-TIME DASHBOARD", 6)
        ws.freeze_panes = "B3"

        closing = self.fx.closing_rates()
        avg     = self.fx.average_rates()
        opening = self.fx.year_opening_rates()

        r = 2
        ws.cell(row=r, column=1).value = "Last Updated:"
        ws.cell(row=r, column=1).font  = self.font_bold
        ws.cell(row=r, column=2).value = datetime.datetime.now().strftime("%Y-%m-%d  %H:%M:%S")
        ws.cell(row=r, column=3).value = "Fiscal Year:"
        ws.cell(row=r, column=4).value = str(_today.year)
        ws.cell(row=r, column=5).value = "FX Source:"
        ws.cell(row=r, column=6).value = "Live API (frankfurter.dev)" if self.fx.is_live() else "Simulated (fallback)"

        # FX Rates section
        r = 4
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1)
        c.value = "FOREIGN EXCHANGE RATES  (EUR BASE)"; c.fill = self.fill_green; c.font = self.font_white_bold; c.alignment = self.align_center

        r += 1
        for hdr, col in [("Currency", 1), ("Closing Rate", 2), ("Year-Avg Rate", 3), ("Year-Opening Rate", 4), ("Change vs Opening", 5), ("Standards Ref", 6)]:
            self._hdr(ws.cell(row=r, column=col), hdr)

        for ccy in FOREIGN_CURRENCIES:
            r += 1
            close_r = closing[ccy]
            open_r  = opening[ccy]
            chg     = _d4((close_r - open_r) / open_r * 100)
            self._data(ws.cell(row=r, column=1), ccy, bold=True)
            self._data(ws.cell(row=r, column=2), float(close_r), "0.0000")
            self._data(ws.cell(row=r, column=3), float(avg[ccy]), "0.0000")
            self._data(ws.cell(row=r, column=4), float(open_r), "0.0000")
            self._data(ws.cell(row=r, column=5), float(chg), "0.00%")
            ws.cell(row=r, column=6).value = "IAS 21 / IFRS 21"
            self._alt_row(ws, r, 6)

        # Balance Sheet
        r += 2
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1)
        c.value = "BALANCE SHEET SUMMARY  (EUR)"; c.fill = self.fill_blue; c.font = self.font_white_bold; c.alignment = self.align_center

        r += 1
        total_a = self.ledger.total_assets()
        total_l = self.ledger.total_liabilities()
        net_eq  = total_a - total_l
        lease_l = sum(ls.current_liability() for ls in self.leases)
        rou_a   = sum(ls.current_rou()       for ls in self.leases)

        for lbl, val, fmt in [
            ("Total Assets (EUR)", float(total_a), "#,##0.00"),
            ("Total Liabilities (EUR)", float(total_l), "#,##0.00"),
            ("Net Equity (EUR)", float(net_eq), "#,##0.00"),
            ("IFRS 16 Lease Liability", float(lease_l), "#,##0.00"),
            ("Right-of-Use Assets", float(rou_a), "#,##0.00"),
        ]:
            ws.cell(row=r, column=1).value = lbl
            ws.cell(row=r, column=1).font  = self.font_bold
            ws.cell(row=r, column=2).value = val
            ws.cell(row=r, column=2).number_format = fmt
            ws.cell(row=r, column=2).font  = self.font_normal
            if val < 0:
                ws.cell(row=r, column=2).fill = self.fill_red
            r += 1

        # Income Statement
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1)
        c.value = "INCOME STATEMENT SUMMARY  (EUR)"; c.fill = self.fill_green; c.font = self.font_white_bold; c.alignment = self.align_center

        r += 1
        rev = self.ledger.total_revenue()
        exp = self.ledger.total_expenses()
        ni  = self.ledger.net_income()
        for lbl, val in [("Total Revenue", float(rev)), ("Total Expenses", float(exp)), ("Net Income / (Loss)", float(ni))]:
            ws.cell(row=r, column=1).value = lbl
            ws.cell(row=r, column=1).font  = self.font_bold
            ws.cell(row=r, column=2).value = val
            ws.cell(row=r, column=2).number_format = "#,##0.00"
            ws.cell(row=r, column=2).fill = self.fill_green_l if val >= 0 else self.fill_red
            r += 1

        # Key Ratios
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1)
        c.value = "KEY RATIOS & FX EXPOSURE"; c.fill = self.fill_purple; c.font = self.font_white_bold; c.alignment = self.align_center

        r += 1
        de_ratio = float(total_l / total_a) if total_a != 0 else 0.0
        op_margin = float(ni / rev) if rev != 0 else 0.0
        ws.cell(row=r, column=1).value = "Debt / Asset Ratio"
        ws.cell(row=r, column=2).value = de_ratio
        ws.cell(row=r, column=2).number_format = "0.00"
        r += 1
        ws.cell(row=r, column=1).value = "Operating Margin"
        ws.cell(row=r, column=2).value = op_margin
        ws.cell(row=r, column=2).number_format = "0.00%"
        r += 1

        mon_fcy = self.ledger.monetary_balances_fcy()
        ws.cell(row=r, column=1).value = "FX Exposure by Currency (FCY net):"
        ws.cell(row=r, column=1).font  = self.font_bold
        r += 1
        for ccy in FOREIGN_CURRENCIES:
            acct_map = mon_fcy.get(ccy, {})
            net_fcy  = sum(acct_map.values()) if acct_map else Decimal("0")
            eur_val  = _d2(net_fcy / closing[ccy]) if closing[ccy] != 0 else Decimal("0")
            ws.cell(row=r, column=1).value = f"  {ccy} exposure"
            ws.cell(row=r, column=2).value = float(net_fcy)
            ws.cell(row=r, column=2).number_format = "#,##0.00"
            ws.cell(row=r, column=3).value = f"EUR equiv:"
            ws.cell(row=r, column=4).value = float(eur_val)
            ws.cell(row=r, column=4).number_format = "#,##0.00"
            r += 1

        self._last_updated(ws)
        self._auto_width(ws)

    # -----------------------------------------------------------------------
    # Sheet 2 — General Ledger EUR
    # -----------------------------------------------------------------------
    def _write_general_ledger(self, ws):
        self._title_row(ws, "GENERAL LEDGER — EUR (Base Currency)  |  Standards: IAS 21, IATA AAG, IFRS 16", 11)
        ws.freeze_panes = "A3"

        hdrs = ["Tx ID", "Date", "Description", "Account", "Type",
                "Debit EUR", "Credit EUR", "Running Balance", "Currency", "FX Rate", "Monetary"]
        r = 2
        for i, h in enumerate(hdrs, 1):
            self._hdr(ws.cell(row=r, column=i), h)

        txs = self.ledger.all_transactions()
        balance = Decimal("0")
        for tx in txs:
            r += 1
            balance += tx.net_eur
            row_vals = [
                tx.tx_id, tx.date.strftime("%d-%b-%Y"), tx.description, tx.account,
                tx.account_type,
                float(tx.debit_eur)  if tx.debit_eur  else "",
                float(tx.credit_eur) if tx.credit_eur else "",
                float(balance),
                tx.currency,
                float(tx.fx_rate),
                "Yes" if tx.is_monetary else "No",
            ]
            self._alt_row(ws, r, 11)
            for col, val in enumerate(row_vals, 1):
                cell = ws.cell(row=r, column=col)
                fmt_map = {6: "#,##0.00", 7: "#,##0.00", 8: "#,##0.00", 10: "0.0000"}
                self._data(cell, val, fmt_map.get(col))

        # Totals
        r += 1
        ws.cell(row=r, column=1).value = "TOTALS"
        ws.cell(row=r, column=1).font  = self.font_bold
        total_dr = sum(t.debit_eur  for t in txs)
        total_cr = sum(t.credit_eur for t in txs)
        for col, val in [(6, float(total_dr)), (7, float(total_cr)), (8, float(total_dr - total_cr))]:
            c = ws.cell(row=r, column=col)
            c.value = val; c.font = self.font_bold; c.number_format = "#,##0.00"
            c.fill = self.fill_alt1
        ws.cell(row=r, column=1).fill = self.fill_alt1

        # Conditional: red on negative running balance
        if len(txs) > 0:
            rng = f"H3:H{r - 1}"
            ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan",  formula=["0"], fill=self.fill_red))
            ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], fill=self.fill_green_l))

        self._last_updated(ws)
        self._auto_width(ws)

    # -----------------------------------------------------------------------
    # Sheets 3–6 — Parallel Ledgers
    # -----------------------------------------------------------------------
    def _write_parallel_ledger(self, ws, ccy: str):
        avg_rates = self.fx.average_rates()
        closing   = self.fx.closing_rates()
        avg_r     = avg_rates.get(ccy, Decimal("1"))
        title = (f"PARALLEL LEDGER — {ccy}  |  IAS 21: transaction-date rate (own currency) "
                 f"or average rate (translated)  |  Ind AS 11")
        self._title_row(ws, title, 12)
        ws.freeze_panes = "A3"

        hdrs = ["Tx ID", "Date", "Description", "Account", "Type",
                f"Debit {ccy}", f"Credit {ccy}", "Running Balance",
                "Rate Used", "Rate Type", "EUR Equiv (closing)", "IAS 21 Ref"]
        r = 2
        for i, h in enumerate(hdrs, 1):
            self._hdr(ws.cell(row=r, column=i))
            ws.cell(row=r, column=i).value = h

        txs = self.ledger.all_transactions()
        balance = Decimal("0")
        for tx in txs:
            r += 1
            if tx.currency == ccy:
                rate_used = tx.fx_rate
                rate_type = "Transaction-Date Rate"
                ias_ref   = "IAS 21 §21"
            else:
                # Income/expense items → average rate; balance sheet monetary → closing rate
                if tx.account_type in ("REVENUE", "EXPENSE"):
                    rate_used = avg_r
                    rate_type = "Period Average Rate"
                    ias_ref   = "IAS 21 §40"
                else:
                    rate_used = closing.get(ccy, avg_r)
                    rate_type = "Closing Rate"
                    ias_ref   = "IAS 21 §23"

            debit_fcy  = _d2(tx.debit_eur  * rate_used) if tx.debit_eur  else Decimal("0")
            credit_fcy = _d2(tx.credit_eur * rate_used) if tx.credit_eur else Decimal("0")
            balance   += debit_fcy - credit_fcy
            eur_equiv  = _d2(balance / closing.get(ccy, Decimal("1")))

            self._alt_row(ws, r, 12)
            vals = [
                tx.tx_id, tx.date.strftime("%d-%b-%Y"), tx.description, tx.account,
                tx.account_type,
                float(debit_fcy)  if debit_fcy  else "",
                float(credit_fcy) if credit_fcy else "",
                float(balance),
                float(rate_used),
                rate_type,
                float(eur_equiv),
                ias_ref,
            ]
            fmt_map = {6: "#,##0.00", 7: "#,##0.00", 8: "#,##0.00", 9: "0.0000", 11: "#,##0.00"}
            for col, val in enumerate(vals, 1):
                self._data(ws.cell(row=r, column=col), val, fmt_map.get(col))

        self._last_updated(ws)
        self._auto_width(ws)

    # -----------------------------------------------------------------------
    # Sheet 7 — Forex Revaluation
    # -----------------------------------------------------------------------
    def _write_forex_revaluation(self, ws):
        closing = self.fx.closing_rates()
        self._title_row(ws, "FOREX REVALUATION — IAS 21 / IFRS 21 / Ind AS 11  |  Closing Rate Method for Monetary Items", 10)
        ws.freeze_panes = "A3"

        hdrs = ["Account", "Currency", "FCY Net Balance", "Booking Rate",
                "EUR at Booking", "Closing Rate", "EUR at Closing",
                "Exchange Diff (EUR)", "Classification", "P&L / OCI Impact"]
        r = 2
        for i, h in enumerate(hdrs, 1):
            self._hdr(ws.cell(row=r, column=i))
            ws.cell(row=r, column=i).value = h

        mon_fcy = self.ledger.monetary_balances_fcy()
        total_gain = Decimal("0")
        total_loss = Decimal("0")
        r_start = 3

        for ccy in FOREIGN_CURRENCIES:
            acct_map = mon_fcy.get(ccy, {})
            for acct, fcy_bal in acct_map.items():
                if fcy_bal == 0:
                    continue

                # Weighted average booking rate
                txs_for = [t for t in self.ledger.all_transactions()
                           if t.is_monetary and t.currency == ccy and t.account == acct]
                sum_fcy = sum(t.debit_fcy + t.credit_fcy for t in txs_for)
                sum_eur = sum(t.debit_eur + t.credit_eur for t in txs_for)
                if sum_eur and sum_eur != 0:
                    booking_rate = _d4(sum_fcy / sum_eur)
                else:
                    booking_rate = closing.get(ccy, Decimal("1"))

                close_r      = closing.get(ccy, Decimal("1"))
                eur_booking  = _d2(fcy_bal / booking_rate)  if booking_rate != 0 else Decimal("0")
                eur_closing  = _d2(fcy_bal / close_r)       if close_r      != 0 else Decimal("0")
                ex_diff      = _d2(eur_closing - eur_booking)
                atype        = _account_type(acct)
                classification = "Monetary" if atype in ("ASSET", "LIABILITY") else "Equity"
                pl_oci         = "P&L" if classification == "Monetary" else "OCI"

                if ex_diff > 0:
                    total_gain += ex_diff
                else:
                    total_loss += ex_diff

                r += 1
                self._alt_row(ws, r, 10)
                vals = [acct, ccy, float(fcy_bal), float(booking_rate), float(eur_booking),
                        float(close_r), float(eur_closing), float(ex_diff), classification, pl_oci]
                fmt_map = {3: "#,##0.00", 4: "0.0000", 5: "#,##0.00", 6: "0.0000",
                           7: "#,##0.00", 8: "#,##0.00"}
                for col, val in enumerate(vals, 1):
                    self._data(ws.cell(row=r, column=col), val, fmt_map.get(col))
                if ex_diff > 0:
                    ws.cell(row=r, column=8).fill = self.fill_green_l
                elif ex_diff < 0:
                    ws.cell(row=r, column=8).fill = self.fill_red

        # Summary
        r += 2
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        c = ws.cell(row=r, column=1)
        c.value = "SUMMARY — UNREALISED FOREIGN EXCHANGE DIFFERENCES  (IAS 21 §28: taken to P&L)"
        c.fill = self.fill_green; c.font = self.font_white_bold; c.alignment = self.align_center

        for lbl, val, clr in [
            ("Total Unrealised Gain (EUR)", float(total_gain), self.fill_green_l),
            ("Total Unrealised Loss (EUR)", float(total_loss), self.fill_red),
            ("Net Exchange Difference (EUR)", float(total_gain + total_loss), None),
        ]:
            r += 1
            ws.cell(row=r, column=1).value = lbl
            ws.cell(row=r, column=1).font  = self.font_bold
            ws.cell(row=r, column=2).value = val
            ws.cell(row=r, column=2).number_format = "#,##0.00"
            ws.cell(row=r, column=2).font = self.font_bold
            if clr:
                ws.cell(row=r, column=2).fill = clr

        r += 2
        note = ("Note: Monetary items (cash, receivables, payables, borrowings) translated at closing rate per IAS 21 §23 / "
                "IFRS 21 §28 / Ind AS 11. Exchange differences recognised in profit or loss. Non-monetary items carried "
                "at historical cost are NOT retranslated (IAS 21 §23(b)).")
        ws.cell(row=r, column=1).value = note
        ws.cell(row=r, column=1).font  = self.font_small
        ws.cell(row=r, column=1).alignment = self.align_wrap
        ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=10)
        ws.row_dimensions[r].height = 28

        self._last_updated(ws)
        self._auto_width(ws)

    # -----------------------------------------------------------------------
    # Sheet 8 — Interest Calculation
    # -----------------------------------------------------------------------
    def _write_interest_calculation(self, ws):
        self._title_row(ws, "INTEREST CALCULATION  |  IFRS 16 EIR Method + Average-Rate Method  |  IAS 23 Borrowing Costs", 10)
        ws.freeze_panes = "A3"
        today = datetime.date.today()
        days_elapsed = max(1, (today - FISCAL_YEAR_START).days)

        # Section A — Average Rate Method
        r = 2
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        c = ws.cell(row=r, column=1)
        c.value = "A.  INTEREST EXPENSE — AVERAGE-OF-YEAR RATE METHOD  (IAS 21 §40 / Ind AS 11)"
        c.fill = self.fill_blue; c.font = self.font_white_bold; c.alignment = self.align_center

        r += 1
        hdrs = ["Instrument / Account", "Principal (EUR)", "Annual Rate %", "Days Elapsed",
                "Days in Year", "Interest (EUR)", "Avg FX Rate", "Interest (FCY)", "Currency", "Standard"]
        for i, h in enumerate(hdrs, 1):
            self._hdr(ws.cell(row=r, column=i))
            ws.cell(row=r, column=i).value = h

        avg = self.fx.average_rates()
        borrowings = [
            ("Term Loan A (Aircraft Finance)", Decimal("10000000"), Decimal("0.0525"), "EUR"),
            ("Revolving Credit Facility",      Decimal("5000000"),  Decimal("0.0450"), "USD"),
            ("Working Capital Facility",       Decimal("2000000"),  Decimal("0.0380"), "GBP"),
        ]
        for name, principal, rate, ccy in borrowings:
            r += 1
            interest_eur = _d2(principal * rate * Decimal(str(days_elapsed)) / Decimal(str(DAYS_IN_YEAR)))
            if ccy == "EUR":
                avg_r      = Decimal("1")
                int_fcy    = interest_eur
            else:
                avg_r      = avg.get(ccy, Decimal("1"))
                int_fcy    = _d2(interest_eur * avg_r)
            self._alt_row(ws, r, 10)
            vals = [name, float(principal), float(rate), days_elapsed, DAYS_IN_YEAR,
                    float(interest_eur), float(avg_r), float(int_fcy), ccy, "IAS 23 / IFRS 9"]
            fmt_map = {2: "#,##0.00", 3: "0.00%", 6: "#,##0.00", 7: "0.0000", 8: "#,##0.00"}
            for col, val in enumerate(vals, 1):
                self._data(ws.cell(row=r, column=col), val, fmt_map.get(col))

        # Section B — IFRS 16 EIR Interest
        r += 2
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        c = ws.cell(row=r, column=1)
        c.value = "B.  IFRS 16 / Ind AS 116 — EFFECTIVE INTEREST RATE METHOD (Current Period)"
        c.fill = self.fill_green; c.font = self.font_white_bold; c.alignment = self.align_center

        r += 1
        hdrs16 = ["Lease ID", "Aircraft", "Period No.", "Payment Date",
                  "Opening Liability", "Interest Charge", "Payment", "Principal", "Closing Liability", "YTD Interest"]
        for i, h in enumerate(hdrs16, 1):
            self._hdr(ws.cell(row=r, column=i))
            ws.cell(row=r, column=i).value = h

        for ls in self.leases:
            sched = ls.compute_schedule()
            idx   = ls.current_period_idx()
            row_d = sched[idx]
            r += 1
            self._alt_row(ws, r, 10)
            vals = [ls.config.lease_id, ls.config.aircraft_type, row_d["period"],
                    row_d["date"].strftime("%d-%b-%Y"),
                    float(row_d["opening_liability"]), float(row_d["interest_charge"]),
                    float(row_d["payment"]),           float(row_d["principal"]),
                    float(row_d["closing_liability"]), float(ls.cumulative_interest_ytd())]
            fmt_map = {5: "#,##0.00", 6: "#,##0.00", 7: "#,##0.00", 8: "#,##0.00", 9: "#,##0.00", 10: "#,##0.00"}
            for col, val in enumerate(vals, 1):
                self._data(ws.cell(row=r, column=col), val, fmt_map.get(col))
            ws.cell(row=r, column=5).fill = self.fill_amber

        # Section C — Formula note
        r += 2
        note = ("Interest calculation methodology:\n"
                "  A. Borrowing costs:  I = P × r × (d / D)   where P=principal, r=annual rate, d=days elapsed, D=days in year\n"
                "  B. Lease interest:   I = Opening Liability × [(1 + r_annual)^(1/12) − 1]   (compound monthly effective rate)\n"
                "Average-of-year FX rate applied to translate interest expense per IAS 21 §40 / Ind AS 11.")
        ws.cell(row=r, column=1).value = note
        ws.cell(row=r, column=1).font  = self.font_small
        ws.cell(row=r, column=1).alignment = self.align_wrap
        ws.merge_cells(start_row=r, start_column=1, end_row=r + 3, end_column=10)
        ws.row_dimensions[r].height = 60

        self._last_updated(ws)
        self._auto_width(ws)

    # -----------------------------------------------------------------------
    # Sheet 9 — IFRS 16 Lease Schedule
    # -----------------------------------------------------------------------
    def _write_ifrs16_schedule(self, ws):
        self._title_row(ws, "IFRS 16 / Ind AS 116 — RIGHT-OF-USE ASSET & LEASE LIABILITY AMORTISATION SCHEDULE", 13)
        ws.freeze_panes = "A3"
        today = datetime.date.today()
        r = 1

        for ls in self.leases:
            cfg = ls.config
            sched = ls.compute_schedule()
            pv    = ls.initial_pv()
            mr    = ls.monthly_rate()

            # Lease header
            r += 2
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
            c = ws.cell(row=r, column=1)
            c.value = (f"Lease: {cfg.lease_id}  |  {cfg.aircraft_type}  |  Commencement: "
                       f"{cfg.commencement.strftime('%d-%b-%Y')}  |  Term: {cfg.lease_term_months} months  |  "
                       f"Monthly Payment: EUR {float(cfg.monthly_payment):,.0f}  |  "
                       f"Effective Rate: {float(cfg.effective_rate):.2%}  |  "
                       f"Initial ROU/Liability: EUR {float(pv):,.0f}  |  Monthly Rate: {float(mr):.6f}")
            c.fill = self.fill_purple; c.font = self.font_white_bold; c.alignment = self.align_wrap
            ws.row_dimensions[r].height = 30

            r += 1
            hdrs = ["Period", "Payment Date", "Opening Liability", "Interest Charge",
                    "Payment", "Principal", "Closing Liability",
                    "ROU Opening", "Depreciation", "ROU Closing",
                    "Cum. Interest", "Cum. Depreciation", "Note"]
            for i, h in enumerate(hdrs, 1):
                self._hdr(ws.cell(row=r, column=i))
                ws.cell(row=r, column=i).value = h

            current_idx = ls.current_period_idx()
            for row_d in sched:
                r += 1
                is_current = (row_d["period"] == sched[current_idx]["period"])
                self._alt_row(ws, r, 13)
                vals = [
                    row_d["period"],
                    row_d["date"].strftime("%d-%b-%Y"),
                    float(row_d["opening_liability"]),
                    float(row_d["interest_charge"]),
                    float(row_d["payment"]),
                    float(row_d["principal"]),
                    float(row_d["closing_liability"]),
                    float(row_d["rou_opening"]),
                    float(row_d["rou_depreciation"]),
                    float(row_d["rou_closing"]),
                    float(row_d["cum_interest"]),
                    float(row_d["cum_depreciation"]),
                    "◀ CURRENT" if is_current else "",
                ]
                fmt_map = {3: "#,##0.00", 4: "#,##0.00", 5: "#,##0.00", 6: "#,##0.00",
                           7: "#,##0.00", 8: "#,##0.00", 9: "#,##0.00", 10: "#,##0.00",
                           11: "#,##0.00", 12: "#,##0.00"}
                for col, val in enumerate(vals, 1):
                    cell = ws.cell(row=r, column=col)
                    self._data(cell, val, fmt_map.get(col))
                    if is_current:
                        cell.fill = self.fill_amber
                        cell.font = self.font_bold

            # Totals
            r += 1
            total_int = sum(d["interest_charge"] for d in sched)
            total_pmt = sum(d["payment"]         for d in sched)
            total_pri = sum(d["principal"]        for d in sched)
            total_dep = sum(d["rou_depreciation"] for d in sched)
            ws.cell(row=r, column=1).value = "TOTALS"
            ws.cell(row=r, column=1).font  = self.font_bold
            ws.cell(row=r, column=1).fill  = self.fill_alt1
            for col, val in [(4, total_int), (5, total_pmt), (6, total_pri), (9, total_dep)]:
                c = ws.cell(row=r, column=col)
                c.value = float(val); c.font = self.font_bold
                c.number_format = "#,##0.00"; c.fill = self.fill_alt1

            r += 1
            note = "Ind AS 116 (India) applies identical mechanics to IFRS 16. Differences exist only in transition provisions (modified retrospective). Figures above apply under both standards."
            ws.cell(row=r, column=1).value = note
            ws.cell(row=r, column=1).font  = self.font_small
            ws.cell(row=r, column=1).alignment = self.align_wrap
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
            ws.row_dimensions[r].height = 20

        self._last_updated(ws)
        self._auto_width(ws)

    # -----------------------------------------------------------------------
    # Sheet 10 — Air India A320 Lease Summary (Ind AS 116 / IFRS 16)
    # -----------------------------------------------------------------------
    def _write_airindia_a320_lease(self, ws):
        closing = self.fx.closing_rates()
        usd_eur = closing.get("USD", Decimal("1.0823"))
        inr_eur = closing.get("INR", Decimal("90.12"))
        today   = datetime.date.today()

        ls  = IFRS16LeaseSchedule(AI_A320_LEASE)
        cfg = AI_A320_LEASE
        sched = ls.compute_schedule()
        pv    = ls.initial_pv()            # in USD (lease currency)
        mr    = ls.monthly_rate()
        pv_eur = _d2(pv / usd_eur)
        pv_inr = _d2(pv * inr_eur / usd_eur)

        curr_idx  = ls.current_period_idx()
        curr_row  = sched[curr_idx]
        curr_liab_usd = curr_row["closing_liability"]
        curr_rou_usd  = curr_row["rou_closing"]
        curr_liab_eur = _d2(curr_liab_usd / usd_eur)
        curr_liab_inr = _d2(curr_liab_usd * inr_eur / usd_eur)
        curr_rou_eur  = _d2(curr_rou_usd  / usd_eur)
        curr_rou_inr  = _d2(curr_rou_usd  * inr_eur / usd_eur)
        cum_int_usd   = ls.cumulative_interest_ytd()
        cum_int_eur   = _d2(cum_int_usd / usd_eur)
        cum_int_inr   = _d2(cum_int_usd * inr_eur / usd_eur)

        # Months elapsed and remaining
        months_elapsed = max(0, (today.year - cfg.commencement.year) * 12
                              + (today.month - cfg.commencement.month))
        months_remaining = max(0, cfg.lease_term_months - months_elapsed)
        lease_end = cfg.commencement + datetime.timedelta(days=30 * cfg.lease_term_months)

        # ── Title ──────────────────────────────────────────────────────────
        self._title_row(ws,
            "AIR INDIA — AIRBUS A320neo LEASE LIABILITY & RIGHT-OF-USE ASSET  |  Ind AS 116 / IFRS 16",
            14)
        ws.freeze_panes = "A4"

        # ── Section A: Lease Parameters ────────────────────────────────────
        r = 2
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
        c = ws.cell(row=r, column=1)
        c.value = "A.  LEASE PARAMETERS & KEY TERMS"
        c.fill = self.fill_blue; c.font = self.font_white_bold; c.alignment = self.align_center

        params = [
            ("Airline",                   "Air India Limited (Tata Group)"),
            ("Aircraft Type",             "Airbus A320neo (CFM LEAP-1A engines)"),
            ("Lease ID",                  cfg.lease_id),
            ("Lease Currency",            "USD (converted to EUR & INR for reporting)"),
            ("Commencement Date",         cfg.commencement.strftime("%d-%b-%Y")),
            ("Lease Term",                f"{cfg.lease_term_months} months (12 years)"),
            ("Lease End Date",            lease_end.strftime("%d-%b-%Y")),
            ("Months Elapsed",            f"{months_elapsed} months"),
            ("Months Remaining",          f"{months_remaining} months"),
            ("Monthly Lease Rental",      f"USD {float(cfg.monthly_payment):,.0f}"),
            ("Incremental Borrowing Rate","8.50% p.a. (Ind AS 116 §26 — BB-rated Indian carrier)"),
            ("Monthly Effective Rate",    f"{float(mr):.6f}  [{float(mr)*100:.4f}%]"),
            ("Initial ROU Asset (USD)",   f"USD {float(pv):>18,.0f}"),
            ("Initial ROU Asset (EUR)",   f"EUR {float(pv_eur):>18,.0f}"),
            ("Initial ROU Asset (INR)",   f"INR {float(pv_inr):>18,.0f}"),
            ("Residual Value Guaranteed", "Nil (operating lease structure)"),
            ("Depreciation Method",       "Straight-line over lease term (Ind AS 116 §31)"),
            ("Monthly Depreciation (USD)",f"USD {float(pv / Decimal(str(cfg.lease_term_months))):>14,.0f}"),
            ("Accounting Standard",       "Ind AS 116 (India) / IFRS 16 (IFRS)"),
            ("Maintenance Reserve",       "USD 50,000/month — excluded from lease liability (variable)"),
            ("FX Translation",            f"USD/EUR: {float(usd_eur):.4f}  |  USD/INR equiv: {float(inr_eur/usd_eur):.2f}"),
        ]

        r += 1
        for lbl, val in params:
            self._alt_row(ws, r, 4)
            ws.cell(row=r, column=1).value = lbl
            ws.cell(row=r, column=1).font  = self.font_bold
            ws.cell(row=r, column=1).border = self.border_thin
            ws.cell(row=r, column=2).value = val
            ws.cell(row=r, column=2).font  = self.font_normal
            ws.cell(row=r, column=2).border = self.border_thin
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            r += 1

        # ── Section B: Current Snapshot ────────────────────────────────────
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
        c = ws.cell(row=r, column=1)
        c.value = (f"B.  CURRENT POSITION AS AT {today.strftime('%d-%b-%Y')}  "
                   f"(Period {curr_row['period']} of {cfg.lease_term_months})")
        c.fill = self.fill_green; c.font = self.font_white_bold; c.alignment = self.align_center

        r += 1
        snap_hdrs = ["", "USD (Lease Currency)", "EUR (Workbook Base)",
                     "INR (Ind AS Reporting)", "Notes"]
        for col, h in enumerate(snap_hdrs, 1):
            self._hdr(ws.cell(row=r, column=col), h,
                      fill=self.fill_blue if col > 1 else self.fill_purple)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=14)

        snapshot = [
            ("Lease Liability (Current Period Close)",
             float(curr_liab_usd), float(curr_liab_eur), float(curr_liab_inr),
             "Present value of remaining payments — IAS 21 closing rate"),
            ("Right-of-Use Asset (Net Book Value)",
             float(curr_rou_usd),  float(curr_rou_eur),  float(curr_rou_inr),
             "Cost less accumulated depreciation — Ind AS 116 §30"),
            ("Difference: ROU − Liability",
             float(curr_rou_usd - curr_liab_usd),
             float(curr_rou_eur - curr_liab_eur),
             float(curr_rou_inr - curr_liab_inr),
             "Positive = ROU > Liability (front-loaded interest effect)"),
            ("Current Month Interest Charge",
             float(curr_row["interest_charge"]),
             float(_d2(curr_row["interest_charge"] / usd_eur)),
             float(_d2(curr_row["interest_charge"] * inr_eur / usd_eur)),
             "EIR method: opening liability × monthly rate"),
            ("Current Month Principal Reduction",
             float(curr_row["principal"]),
             float(_d2(curr_row["principal"] / usd_eur)),
             float(_d2(curr_row["principal"] * inr_eur / usd_eur)),
             "Monthly payment − interest charge"),
            ("Current Month ROU Depreciation",
             float(curr_row["rou_depreciation"]),
             float(_d2(curr_row["rou_depreciation"] / usd_eur)),
             float(_d2(curr_row["rou_depreciation"] * inr_eur / usd_eur)),
             "Straight-line: initial PV ÷ lease term months"),
            ("YTD Interest Expense (this fiscal year)",
             float(cum_int_usd), float(cum_int_eur), float(cum_int_inr),
             "Sum of EIR interest charges Apr–current month"),
            ("Cumulative Interest to Date",
             float(curr_row["cum_interest"]),
             float(_d2(curr_row["cum_interest"] / usd_eur)),
             float(_d2(curr_row["cum_interest"] * inr_eur / usd_eur)),
             "Total finance cost recognised since commencement"),
            ("Cumulative Depreciation to Date",
             float(curr_row["cum_depreciation"]),
             float(_d2(curr_row["cum_depreciation"] / usd_eur)),
             float(_d2(curr_row["cum_depreciation"] * inr_eur / usd_eur)),
             "Total depreciation charged since commencement"),
        ]

        for row_data in snapshot:
            r += 1
            lbl, usd_v, eur_v, inr_v, note = row_data
            self._alt_row(ws, r, 14)
            ws.cell(row=r, column=1).value = lbl
            ws.cell(row=r, column=1).font  = self.font_bold
            ws.cell(row=r, column=1).border = self.border_thin
            for col, val in [(2, usd_v), (3, eur_v), (4, inr_v)]:
                cell = ws.cell(row=r, column=col)
                cell.value = val
                cell.number_format = "#,##0.00"
                cell.font  = self.font_normal
                cell.border = self.border_thin
                cell.alignment = self.align_right
                if val < 0:
                    cell.fill = self.fill_red
            ws.cell(row=r, column=5).value = note
            ws.cell(row=r, column=5).font  = self.font_small
            ws.cell(row=r, column=5).border = self.border_thin
            ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=14)

        # Highlight lease liability and ROU rows
        ws.cell(row=r - 8, column=1).fill = self.fill_amber   # Lease Liability row label
        ws.cell(row=r - 7, column=1).fill = self.fill_amber   # ROU Asset row label

        # ── Section C: Annual roll-forward (next 5 years + totals) ─────────
        r += 2
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
        c = ws.cell(row=r, column=1)
        c.value = "C.  ANNUAL ROLL-FORWARD — LEASE LIABILITY & ROU ASSET  (next 5 years + lifetime totals)"
        c.fill = self.fill_purple; c.font = self.font_white_bold; c.alignment = self.align_center

        r += 1
        ann_hdrs = [
            "Fiscal Year", "Opening Liability (USD)", "Interest (USD)", "Payments (USD)",
            "Closing Liability (USD)", "ROU Opening (USD)", "Depreciation (USD)",
            "ROU Closing (USD)", "Closing Liability (EUR)", "Closing Liability (INR)",
            "ROU Closing (EUR)", "ROU Closing (INR)", "Periods", "Ind AS 116 Para"
        ]
        for col, h in enumerate(ann_hdrs, 1):
            self._hdr(ws.cell(row=r, column=col), h)

        # Build annual aggregates from monthly schedule
        annual: dict = {}
        for row_d in sched:
            yr = row_d["date"].year
            if yr not in annual:
                annual[yr] = {
                    "open_liab": row_d["opening_liability"],
                    "interest":  Decimal("0"), "payments": Decimal("0"),
                    "principal": Decimal("0"), "close_liab": Decimal("0"),
                    "rou_open":  row_d["rou_opening"],
                    "dep":       Decimal("0"), "rou_close": Decimal("0"),
                    "periods":   0,
                }
            annual[yr]["interest"]  += row_d["interest_charge"]
            annual[yr]["payments"]  += row_d["payment"]
            annual[yr]["principal"] += row_d["principal"]
            annual[yr]["close_liab"] = row_d["closing_liability"]
            annual[yr]["dep"]       += row_d["rou_depreciation"]
            annual[yr]["rou_close"]  = row_d["rou_closing"]
            annual[yr]["periods"]   += 1

        shown_years = sorted(annual.keys())
        # Show current year + 4 future years + a lifetime total row
        current_year = today.year
        years_to_show = [y for y in shown_years if y >= current_year][:5]

        total_int = Decimal("0"); total_pmt = Decimal("0"); total_dep = Decimal("0")
        for yr in years_to_show:
            a = annual[yr]
            r += 1
            self._alt_row(ws, r, 14)
            cl_eur = _d2(a["close_liab"] / usd_eur)
            cl_inr = _d2(a["close_liab"] * inr_eur / usd_eur)
            rc_eur = _d2(a["rou_close"]  / usd_eur)
            rc_inr = _d2(a["rou_close"]  * inr_eur / usd_eur)
            is_current = (yr == current_year)
            total_int += a["interest"]; total_pmt += a["payments"]; total_dep += a["dep"]
            vals = [
                f"FY {yr}" + (" ◀ CURRENT" if is_current else ""),
                float(a["open_liab"]),  float(a["interest"]),  float(a["payments"]),
                float(a["close_liab"]), float(a["rou_open"]),  float(a["dep"]),
                float(a["rou_close"]),  float(cl_eur),          float(cl_inr),
                float(rc_eur),          float(rc_inr),          a["periods"],
                "Ind AS 116 §36–§38",
            ]
            fmt_map = {2: "#,##0.00", 3: "#,##0.00", 4: "#,##0.00", 5: "#,##0.00",
                       6: "#,##0.00", 7: "#,##0.00", 8: "#,##0.00", 9: "#,##0.00",
                       10: "#,##0.00", 11: "#,##0.00", 12: "#,##0.00"}
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=r, column=col)
                self._data(cell, val, fmt_map.get(col))
                if is_current:
                    cell.fill = self.fill_amber
                    cell.font = self.font_bold

        # Lifetime totals row
        r += 1
        ws.cell(row=r, column=1).value = "LIFETIME TOTALS (all periods)"
        ws.cell(row=r, column=1).font  = self.font_bold
        ws.cell(row=r, column=1).fill  = self.fill_alt1
        for col, val in [
            (3, float(sum(a["interest"] for a in annual.values()))),
            (4, float(sum(a["payments"] for a in annual.values()))),
            (7, float(sum(a["dep"]      for a in annual.values()))),
        ]:
            c2 = ws.cell(row=r, column=col)
            c2.value = val; c2.font = self.font_bold
            c2.number_format = "#,##0.00"; c2.fill = self.fill_alt1

        # ── Section D: Balance Sheet Presentation ─────────────────────────
        r += 2
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
        c = ws.cell(row=r, column=1)
        c.value = "D.  BALANCE SHEET CLASSIFICATION  (Ind AS 116 §47 / IFRS 16 §47)"
        c.fill = self.fill_blue; c.font = self.font_white_bold; c.alignment = self.align_center

        # Current vs non-current split
        # Current portion = principal payments due in next 12 months
        current_principal = Decimal("0")
        for row_d in sched:
            if today <= row_d["date"] <= today + datetime.timedelta(days=365):
                current_principal += row_d["principal"]
        non_current_liab = max(Decimal("0"), curr_liab_usd - current_principal)

        r += 1
        bs_hdrs = ["Balance Sheet Line Item", "USD", "EUR", "INR",
                   "Classification", "Ind AS 116 Para"]
        for col, h in enumerate(bs_hdrs, 1):
            self._hdr(ws.cell(row=r, column=col), h)
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=14)

        bs_rows = [
            ("Right-of-Use Asset — Airbus A320neo",
             float(curr_rou_usd), float(curr_rou_eur), float(curr_rou_inr),
             "Non-Current Asset", "§47(a)"),
            ("Lease Liability — Current portion (due < 12 months)",
             float(current_principal),
             float(_d2(current_principal / usd_eur)),
             float(_d2(current_principal * inr_eur / usd_eur)),
             "Current Liability", "§47(b)"),
            ("Lease Liability — Non-current portion (due > 12 months)",
             float(non_current_liab),
             float(_d2(non_current_liab / usd_eur)),
             float(_d2(non_current_liab * inr_eur / usd_eur)),
             "Non-Current Liability", "§47(b)"),
            ("Total Lease Liability",
             float(curr_liab_usd), float(curr_liab_eur), float(curr_liab_inr),
             "Total", "§26"),
            ("Interest Expense (YTD — P&L)",
             float(cum_int_usd), float(cum_int_eur), float(cum_int_inr),
             "Finance Costs — P&L", "§49(b)"),
            ("Depreciation Expense (YTD — P&L)",
             float(_d2(curr_row["rou_depreciation"] * Decimal(str(months_elapsed % 12 or 12)))),
             float(_d2(curr_row["rou_depreciation"] * Decimal(str(months_elapsed % 12 or 12)) / usd_eur)),
             float(_d2(curr_row["rou_depreciation"] * Decimal(str(months_elapsed % 12 or 12)) * inr_eur / usd_eur)),
             "Depreciation — P&L", "§49(a)"),
        ]
        for row_data in bs_rows:
            r += 1
            lbl, usd_v, eur_v, inr_v, cls, para = row_data
            self._alt_row(ws, r, 14)
            ws.cell(row=r, column=1).value = lbl
            ws.cell(row=r, column=1).font  = self.font_bold
            ws.cell(row=r, column=1).border = self.border_thin
            for col, val in [(2, usd_v), (3, eur_v), (4, inr_v)]:
                cell = ws.cell(row=r, column=col)
                cell.value = val; cell.number_format = "#,##0.00"
                cell.font = self.font_normal; cell.border = self.border_thin
                cell.alignment = self.align_right
            ws.cell(row=r, column=5).value  = cls
            ws.cell(row=r, column=5).font   = self.font_normal
            ws.cell(row=r, column=5).border = self.border_thin
            ws.cell(row=r, column=6).value  = para
            ws.cell(row=r, column=6).font   = self.font_small
            ws.cell(row=r, column=6).border = self.border_thin
            ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=14)

        # ── Section E: Full Monthly Schedule (all periods) ─────────────────
        r += 2
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
        c = ws.cell(row=r, column=1)
        c.value = "E.  FULL MONTHLY AMORTISATION SCHEDULE (144 periods)"
        c.fill = self.fill_purple; c.font = self.font_white_bold; c.alignment = self.align_center

        r += 1
        full_hdrs = ["Period", "Date", "Opening Liability", "Interest",
                     "Payment", "Principal", "Closing Liability",
                     "ROU Opening", "Depreciation", "ROU Closing",
                     "Cum. Interest", "Cum. Depreciation",
                     "Liab (EUR)", "ROU (EUR)"]
        for col, h in enumerate(full_hdrs, 1):
            self._hdr(ws.cell(row=r, column=col), h)

        for row_d in sched:
            r += 1
            is_curr = (row_d["period"] == curr_row["period"])
            self._alt_row(ws, r, 14)
            liab_eur = float(_d2(row_d["closing_liability"] / usd_eur))
            rou_eur  = float(_d2(row_d["rou_closing"]       / usd_eur))
            vals = [
                row_d["period"],
                row_d["date"].strftime("%d-%b-%Y"),
                float(row_d["opening_liability"]),
                float(row_d["interest_charge"]),
                float(row_d["payment"]),
                float(row_d["principal"]),
                float(row_d["closing_liability"]),
                float(row_d["rou_opening"]),
                float(row_d["rou_depreciation"]),
                float(row_d["rou_closing"]),
                float(row_d["cum_interest"]),
                float(row_d["cum_depreciation"]),
                liab_eur, rou_eur,
            ]
            fmt_map = {3: "#,##0.00", 4: "#,##0.00", 5: "#,##0.00", 6: "#,##0.00",
                       7: "#,##0.00", 8: "#,##0.00", 9: "#,##0.00", 10: "#,##0.00",
                       11: "#,##0.00", 12: "#,##0.00", 13: "#,##0.00", 14: "#,##0.00"}
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=r, column=col)
                self._data(cell, val, fmt_map.get(col))
                if is_curr:
                    cell.fill = self.fill_amber
                    cell.font = self.font_bold

        # ── Disclosure notes ───────────────────────────────────────────────
        r += 2
        notes = [
            "DISCLOSURE NOTES (Ind AS 116 §52–§60 / IFRS 16 §52–§60):",
            f"1. Lease liability measured at present value of {cfg.lease_term_months} monthly payments of USD {float(cfg.monthly_payment):,.0f} discounted at 8.50% p.a. incremental borrowing rate.",
            "2. ROU asset initially measured equal to lease liability (no initial direct costs, lease incentives, or prepayments in this model).",
            "3. ROU asset depreciated straight-line over lease term (12 years). No residual value guaranteed.",
            "4. Lease liability retranslated at closing USD/INR rate at each reporting date per Ind AS 11 / IAS 21. Exchange differences to P&L.",
            "5. Variable lease payments (maintenance reserves ~USD 50,000/month) excluded from lease liability — expensed as incurred per Ind AS 116 §38(b).",
            "6. Air India adopted Ind AS 116 from FY2019-20 under modified retrospective approach; new leases recognised at commencement date.",
            "7. Current/non-current split based on contractual undiscounted cash flows due within / after 12 months of reporting date.",
            f"8. Market data basis: AVAC / IBA narrow-body lease rate survey 2023-24. IBR derived from Air India BB credit profile + Indian sovereign rate.",
        ]
        for note in notes:
            ws.cell(row=r, column=1).value = note
            ws.cell(row=r, column=1).font  = self.font_small if note.startswith("  ") or note[0].isdigit() else Font(name="Calibri", bold=True, size=9)
            ws.cell(row=r, column=1).alignment = self.align_wrap
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
            ws.row_dimensions[r].height = 16
            r += 1

        self._last_updated(ws)

        # Manual column widths for this sheet
        col_widths = [42, 18, 18, 20, 18, 18, 18, 18, 18, 18, 18, 18, 18, 18]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # -----------------------------------------------------------------------
    # Sheet 11 — Year-End Closing
    # -----------------------------------------------------------------------
    def _write_year_end_closing(self, ws):
        closing = self.fx.closing_rates()
        opening = self.fx.year_opening_rates()
        self._title_row(ws, "YEAR-END CLOSING ENTRIES — IAS 21 §23 / §28 / §30  |  Closing Rate Revaluation of Monetary Items", 11)
        ws.freeze_panes = "A3"

        # Section A
        r = 2
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
        c = ws.cell(row=r, column=1)
        c.value = "A.  CLOSING RATE REVALUATION OF ALL MONETARY LIABILITIES & ASSETS"
        c.fill = self.fill_blue; c.font = self.font_white_bold; c.alignment = self.align_center

        r += 1
        hdrs = ["Account", "Currency", "FCY Balance", "Year-Opening Rate",
                "Year-Closing Rate", "EUR at Opening", "EUR at Closing",
                "Exchange Gain/(Loss)", "Status", "Treatment", "IAS 21 Para"]
        for i, h in enumerate(hdrs, 1):
            self._hdr(ws.cell(row=r, column=i))
            ws.cell(row=r, column=i).value = h

        mon_fcy = self.ledger.monetary_balances_fcy()
        total_gain_r = Decimal("0")
        total_loss_r = Decimal("0")
        total_gain_u = Decimal("0")
        total_loss_u = Decimal("0")

        for ccy in FOREIGN_CURRENCIES:
            acct_map = mon_fcy.get(ccy, {})
            for acct, fcy_bal in acct_map.items():
                if fcy_bal == 0:
                    continue
                open_r  = opening.get(ccy, Decimal("1"))
                close_r = closing.get(ccy, Decimal("1"))
                eur_open  = _d2(fcy_bal / open_r)  if open_r  != 0 else Decimal("0")
                eur_close = _d2(fcy_bal / close_r) if close_r != 0 else Decimal("0")
                ex_diff   = _d2(eur_close - eur_open)
                # Determine realised vs unrealised based on account type
                atype = _account_type(acct)
                status    = "Unrealised" if atype == "LIABILITY" else "Realised"
                treatment = "P&L" if atype in ("ASSET", "LIABILITY") else "OCI"
                ias_para  = "IAS 21 §28" if treatment == "P&L" else "IAS 21 §32"

                if status == "Realised":
                    if ex_diff >= 0: total_gain_r += ex_diff
                    else:            total_loss_r += ex_diff
                else:
                    if ex_diff >= 0: total_gain_u += ex_diff
                    else:            total_loss_u += ex_diff

                r += 1
                self._alt_row(ws, r, 11)
                vals = [acct, ccy, float(fcy_bal), float(open_r), float(close_r),
                        float(eur_open), float(eur_close), float(ex_diff), status, treatment, ias_para]
                fmt_map = {3: "#,##0.00", 4: "0.0000", 5: "0.0000",
                           6: "#,##0.00", 7: "#,##0.00", 8: "#,##0.00"}
                for col, val in enumerate(vals, 1):
                    self._data(ws.cell(row=r, column=col), val, fmt_map.get(col))
                ws.cell(row=r, column=8).fill = self.fill_green_l if ex_diff >= 0 else self.fill_red

        # Section B — Summary statistics
        r += 2
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
        c = ws.cell(row=r, column=1)
        c.value = "B.  YEAR-END FX SUMMARY"
        c.fill = self.fill_green; c.font = self.font_white_bold; c.alignment = self.align_center

        tax_rate = Decimal("0.25")
        net_pl = (total_gain_r + total_loss_r + total_gain_u + total_loss_u)
        tax_eff = _d2(net_pl * tax_rate)
        post_tax = _d2(net_pl - tax_eff)

        rows_b = [
            ("Realised FX Gain (EUR)",      float(total_gain_r), self.fill_green_l),
            ("Realised FX Loss (EUR)",       float(total_loss_r), self.fill_red),
            ("Unrealised FX Gain (EUR)",     float(total_gain_u), self.fill_green_l),
            ("Unrealised FX Loss (EUR)",     float(total_loss_u), self.fill_red),
            ("Net FX Impact to P&L (EUR)",   float(net_pl),       None),
            ("Assumed Tax Rate",             float(tax_rate),     None),
            ("Tax Effect on FX (EUR)",       float(tax_eff),      self.fill_amber),
            ("Post-Tax FX Impact (EUR)",     float(post_tax),     None),
        ]
        for lbl, val, fill_obj in rows_b:
            r += 1
            ws.cell(row=r, column=1).value = lbl
            ws.cell(row=r, column=1).font  = self.font_bold
            cell_v = ws.cell(row=r, column=2)
            cell_v.value = val
            cell_v.font  = self.font_bold
            cell_v.number_format = "0.00%" if lbl == "Assumed Tax Rate" else "#,##0.00"
            if fill_obj:
                cell_v.fill = fill_obj

        # Section C — Closing journal entries
        r += 2
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
        c = ws.cell(row=r, column=1)
        c.value = "C.  INDICATIVE CLOSING JOURNAL ENTRIES  (IAS 21 §23–§30)"
        c.fill = self.fill_purple; c.font = self.font_white_bold; c.alignment = self.align_center

        r += 1
        for col, h in enumerate(["#", "Dr Account", "Cr Account", "Amount (EUR)", "Narrative"], 1):
            self._hdr(ws.cell(row=r, column=col))
            ws.cell(row=r, column=col).value = h

        jnls = [
            (1, "Foreign Currency Monetary Liabilities",  "FX Gain (P&L)",             float(max(Decimal("0"), total_gain_r + total_gain_u)), "Upward revaluation at closing rate"),
            (2, "FX Loss (P&L)",  "Foreign Currency Monetary Assets",   float(abs(min(Decimal("0"), total_loss_r + total_loss_u))), "Downward revaluation at closing rate"),
            (3, "FX P&L Summary", "Retained Earnings",                   float(abs(net_pl)), "Close net FX P&L to equity"),
        ]
        for jnl in jnls:
            r += 1
            self._alt_row(ws, r, 5)
            for col, val in enumerate(jnl, 1):
                cell = ws.cell(row=r, column=col)
                cell.value = val
                cell.font  = self.font_normal
                if col == 4:
                    cell.number_format = "#,##0.00"

        self._last_updated(ws)
        self._auto_width(ws)

    # -----------------------------------------------------------------------
    # Sheet 11 — Standards Reference
    # -----------------------------------------------------------------------
    def _write_standards_reference(self, ws):
        self._title_row(ws, "ACCOUNTING STANDARDS APPLIED — AIRLINE ACCOUNTING REFERENCE TABLE", 6)
        ws.freeze_panes = "A3"

        hdrs = ["Standard", "Full Name", "Accounting Treatment", "Measurement Basis",
                "Recognition Criteria", "Sheets Affected"]
        r = 2
        for i, h in enumerate(hdrs, 1):
            self._hdr(ws.cell(row=r, column=i), h)

        standards = [
            ("IAS 21",     "Effects of Changes in Foreign Exchange Rates",
             "Monetary items translated at closing rate; exchange differences recognised in P&L; non-monetary at historical rate",
             "Historical cost (non-monetary) / Closing rate (monetary)",
             "At each reporting date for monetary items; at transaction date for initial recognition",
             "Forex Revaluation, Year-End Closing, Parallel Ledgers"),

            ("IFRS 21",    "Presentation of Financial Statements — Foreign Currency (effective 2025, replaces IAS 21)",
             "Same principles as IAS 21; clearer functional vs presentation currency distinction; updated transition guidance",
             "Functional currency; closing rate for monetary items",
             "At each reporting date",
             "All FX sheets"),

            ("IFRS 16",    "Leases",
             "Lessee recognises ROU asset and lease liability at PV of future payments; interest via effective interest rate; depreciation of ROU asset straight-line",
             "Present value of lease payments (effective interest method for liability; cost model for ROU asset)",
             "At lease commencement date; reassessed on modification",
             "IFRS 16 Lease Schedule, Interest Calculation, General Ledger EUR"),

            ("Ind AS 116", "Leases (India equivalent of IFRS 16)",
             "Identical to IFRS 16 in steady-state; transitional differences (modified retrospective vs full retrospective)",
             "Present value of lease payments",
             "Lease commencement; modified retrospective transition available",
             "IFRS 16 Lease Schedule, Parallel Ledger INR"),

            ("Ind AS 11",  "Construction Contracts / Foreign Exchange Effects (India equivalent of IAS 21)",
             "Monetary items at closing rate; exchange differences to P&L; non-monetary at transaction-date rate",
             "Closing rate (monetary); historical rate (non-monetary)",
             "Reporting date for monetary items",
             "Forex Revaluation, Parallel Ledger INR, Year-End Closing"),

            ("IATA AAG",   "IATA Airline Accounting Guidelines (Industry Standard)",
             "Standardised chart of accounts; passenger revenue recognised on flight completion; cargo on shipment; maintenance reserves as constructive obligations",
             "Accrual basis; agent vs principal distinction for codeshares",
             "Revenue: service delivery (flight completion / cargo shipment). Maintenance: when obligation probable",
             "General Ledger EUR, Dashboard"),

            ("IFRS 9",     "Financial Instruments",
             "Fuel hedges designated as cash flow hedges at fair value through OCI; borrowings at amortised cost; hedge ineffectiveness to P&L",
             "Fair value (derivatives); Amortised cost (borrowings)",
             "When instrument is designated and hedge documentation prepared",
             "General Ledger EUR, Interest Calculation"),

            ("IAS 16",     "Property, Plant and Equipment",
             "Owned aircraft at cost less accumulated depreciation and impairment; component accounting required",
             "Cost model (or optional revaluation model)",
             "When future economic benefits probable and cost measurable",
             "General Ledger EUR"),

            ("IAS 36",     "Impairment of Assets",
             "ROU assets and owned aircraft tested for impairment when indicators exist; carrying amount vs recoverable amount",
             "Recoverable amount = higher of fair value less costs of disposal and value in use",
             "When impairment indicators identified (annual for goodwill)",
             "IFRS 16 Lease Schedule"),

            ("IAS 37",     "Provisions, Contingent Liabilities and Contingent Assets",
             "Maintenance reserves recognised as constructive obligations based on utilisation; discount to present value for long-term provisions",
             "Best estimate of expenditure; discounted if material",
             "When present obligation probable and can be measured reliably",
             "General Ledger EUR"),

            ("IAS 12",     "Income Taxes",
             "Deferred tax recognised on all temporary differences including FX and lease-related differences; IAS 12 §41A exemption for initial recognition of ROU assets/lease liabilities",
             "Liability method at enacted rates",
             "At each reporting date",
             "Year-End Closing, Dashboard"),

            ("IAS 7",      "Statement of Cash Flows",
             "FX effect on cash and cash equivalents disclosed as separate reconciling item; lease payments split into principal (financing) and interest (operating or financing per policy)",
             "Actual cash flows; FX effect shown separately",
             "Each reporting period",
             "Dashboard"),
        ]

        for i, row_data in enumerate(standards):
            r += 1
            fill = self.fill_alt1 if i % 2 == 0 else self.fill_alt2
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=col)
                cell.value = val
                cell.font  = self.font_bold if col == 1 else self.font_normal
                cell.fill  = fill
                cell.alignment = self.align_wrap
                cell.border = self.border_thin
            ws.row_dimensions[r].height = 50

        self._last_updated(ws)
        # Manual widths for this reference sheet
        widths = [12, 42, 55, 32, 42, 28]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # -----------------------------------------------------------------------
    # Sheet 13 — Airline Lease Comparison
    # -----------------------------------------------------------------------
    def _write_airline_lease_comparison(self, ws):
        closing  = self.fx.closing_rates()
        usd_rate = closing.get("USD", Decimal("1.0823"))   # EUR→USD; divide to get EUR value
        today    = datetime.date.today()

        rev  = self.ledger.total_revenue()
        exp  = self.ledger.total_expenses()

        pax_rev = abs(sum(
            t.net_eur for t in self.ledger.all_transactions()
            if t.account == "Passenger Revenue"
        )) or Decimal("1")
        cargo_rev = abs(sum(
            t.net_eur for t in self.ledger.all_transactions()
            if t.account == "Cargo Revenue"
        )) or Decimal("0")
        total_op_rev = pax_rev + cargo_rev or Decimal("1")

        # ── Title ──────────────────────────────────────────────────────────
        self._title_row(ws,
            "AIRLINE LEASE COMPARISON — MULTI-STANDARD ANALYSIS, FLEET PROFITABILITY & MARKET RATES",
            12)
        ws.freeze_panes = "A4"
        r = 1

        # ====================================================================
        # SECTION A — Multi-Airline Accounting Standards Comparison
        # ====================================================================
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
        c = ws.cell(row=r, column=1)
        c.value = "A.  MULTI-AIRLINE LEASE ACCOUNTING STANDARDS COMPARISON  (FY2023/24 Annual Report Basis)"
        c.fill = self.fill_blue; c.font = self.font_white_bold; c.alignment = self.align_center

        r += 1
        a_hdrs = ["Airline", "Jurisdiction", "Standard", "Discount Rate (IBR) %",
                  "Fleet Size", "Leased %", "Lease Liability (USD bn)",
                  "ROU Asset (USD bn)", "Annual Lease Cost (USD m)",
                  "Key Difference vs IFRS 16", "Credit Rating"]
        for i, h in enumerate(a_hdrs, 1):
            self._hdr(ws.cell(row=r, column=i), h)

        AIRLINE_DATA = [
            # (Airline, Jurisdiction, Standard, IBR, Fleet, Leased%, LiabUSDbn, ROUbn, CostUSDm, KeyDiff, Rating)
            ("Air India", "India", "Ind AS 116", 0.085, 220, 0.78, 4.2, 3.9, 340.0,
             "Identical to IFRS 16 mechanically; modified retrospective transition; IBR from Indian BB sovereign spread (~7.2% G-Sec + 130bps)",
             "BB (CRISIL)"),
            ("Lufthansa Group", "Germany / EU", "IFRS 16", 0.026, 763, 0.52, 7.8, 7.1, 1420.0,
             "Reference standard — lessee on-balance-sheet model; EIR method; lowest IBR among peers (EUR IG-rated)",
             "BBB- (S&P)"),
            ("Delta Air Lines", "USA", "ASC 842", 0.031, 1256, 0.61, 11.3, 10.4, 2100.0,
             "ASC 842: operating leases show straight-line P&L (no interest/depreciation split); finance leases like IFRS 16; both on B/S",
             "BB+ (S&P)"),
            ("Singapore Airlines", "Singapore", "IFRS 16 (FRS 116)", 0.034, 213, 0.44, 9.1, 8.6, 1240.0,
             "FRS 116 = verbatim IFRS 16 adoption; strong government-linked balance sheet yields lower IBR vs peers",
             "AA- (Fitch)"),
            ("Emirates", "UAE", "IFRS 16", 0.038, 260, 0.87, 17.4, 16.2, 2890.0,
             "IFRS 16 applied; highest lease% in peer group (all-leased model); no UAE corp tax pre-2023; govt-backed implicit guarantee",
             "A (Moody's)"),
            ("Ryanair", "Ireland / EU", "IFRS 16", 0.027, 569, 0.92, 4.9, 4.6, 680.0,
             "IFRS 16; unusually high lease% but short avg term ~7yr vs industry 12yr; low IBR due to Irish market + IG rating",
             "BBB+ (S&P)"),
            ("IndiGo (InterGlobe)", "India", "Ind AS 116", 0.092, 350, 0.95, 6.1, 5.7, 520.0,
             "Ind AS 116; higher IBR than Air India (BB- vs BB); all-A320 family fleet; dominant sale-leaseback model",
             "BB- (ICRA)"),
            ("Cathay Pacific", "Hong Kong", "HKFRS 16", 0.030, 253, 0.71, 12.8, 11.9, 1650.0,
             "HKFRS 16 = word-for-word IFRS 16 (HKICPA); wide-body heavy fleet inflates liability/aircraft vs narrow-body peers",
             "BBB (S&P)"),
        ]

        for i, row_data in enumerate(AIRLINE_DATA):
            r += 1
            self._alt_row(ws, r, 11)
            (airline, jur, std, ibr, fleet, leased_pct,
             liab, rou, cost, diff, rating) = row_data
            vals = [airline, jur, std, ibr, fleet, leased_pct,
                    liab, rou, cost, diff, rating]
            fmt_map = {4: "0.00%", 6: "0.0%", 7: "#,##0.0", 8: "#,##0.0", 9: "#,##0.0"}
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=r, column=col)
                self._data(cell, val, fmt_map.get(col))
                if col == 10:
                    cell.alignment = self.align_wrap
                if col == 1:
                    cell.font = self.font_bold
            ws.row_dimensions[r].height = 40

        r += 1
        src_note = ("Sources: Annual Reports FY2023-24 — Lufthansa AR2023, Delta 10-K 2023, SIA AR2023/24, "
                    "Emirates AR2023/24, Ryanair AR2024, IndiGo FY24, Cathay Pacific AR2023, Air India FY24. "
                    "IBR = Incremental Borrowing Rate per IFRS 16 §26 / Ind AS 116 §26. "
                    "Lease liabilities in USD billions at closing FX rate.")
        ws.cell(row=r, column=1).value = src_note
        ws.cell(row=r, column=1).font  = self.font_small
        ws.cell(row=r, column=1).alignment = self.align_wrap
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
        ws.row_dimensions[r].height = 24

        # ====================================================================
        # SECTION B — Own Fleet Lease Profitability Analysis
        # ====================================================================
        r += 2
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
        c = ws.cell(row=r, column=1)
        c.value = "B.  OWN FLEET LEASE PROFITABILITY ANALYSIS  (Based on Live Ledger Data — Real-Time)"
        c.fill = self.fill_green; c.font = self.font_white_bold; c.alignment = self.align_center

        r += 1
        b_hdrs = ["Lease ID", "Aircraft Type", "Monthly Rental (EUR)",
                  "Est. Monthly Revenue (EUR)", "Monthly Op. Costs (EUR)",
                  "Net Contribution (EUR)", "Break-Even Load Factor %",
                  "Est. Current Load Factor %", "Profit / (Loss) EUR",
                  "ROI on Lease %", "Status"]
        for i, h in enumerate(b_hdrs, 1):
            self._hdr(ws.cell(row=r, column=i), h, fill=self.fill_green)

        n_ac = len(self.leases) + 1   # self.leases + AI_A320
        lease_exp_total = sum(ls.config.monthly_payment for ls in self.leases) * Decimal("12")
        non_lease_exp   = max(Decimal("0"), exp - lease_exp_total)

        op_margin = float(self.ledger.net_income() / rev) if rev != 0 else 0.05
        est_lf    = min(0.97, max(0.50, 0.75 + (op_margin - 0.05) * 2))

        fleet_items = []
        for ls in self.leases:
            fleet_items.append(("self", ls))
        fleet_items.append(("ai", None))

        for idx, (tag, ls) in enumerate(fleet_items):
            r += 1
            if tag == "self":
                cfg            = ls.config
                monthly_rental = cfg.monthly_payment
                aircraft_type  = cfg.aircraft_type
                lease_id       = cfg.lease_id
                currency_note  = "EUR"
            else:
                monthly_rental = _d2(AI_A320_LEASE.monthly_payment / usd_rate)
                aircraft_type  = "Airbus A320neo (Air India - VT-RTA)"
                lease_id       = AI_A320_LEASE.lease_id
                currency_note  = "EUR (conv. from USD)"

            monthly_rev   = _d2(rev / Decimal("12") / Decimal(str(n_ac)))
            monthly_opcost = _d2(non_lease_exp / Decimal("12") / Decimal(str(n_ac)))
            net_contrib   = _d2(monthly_rev - monthly_rental - monthly_opcost)
            total_cost    = monthly_rental + monthly_opcost
            breakeven_lf  = float(total_cost / monthly_rev) if monthly_rev != 0 else 0.0
            roi           = float(net_contrib / monthly_rental) if monthly_rental != 0 else 0.0
            is_profit     = net_contrib > 0

            self._alt_row(ws, r, 11)
            vals = [lease_id, aircraft_type,
                    float(monthly_rental), float(monthly_rev), float(monthly_opcost),
                    float(net_contrib), breakeven_lf, est_lf,
                    float(net_contrib), roi,
                    "PROFIT" if is_profit else "LOSS"]
            fmt_map = {3: "#,##0", 4: "#,##0", 5: "#,##0", 6: "#,##0",
                       7: "0.0%", 8: "0.0%", 9: "#,##0", 10: "0.0%"}
            profit_fill = self.fill_green_l if is_profit else self.fill_red
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=r, column=col)
                self._data(cell, val, fmt_map.get(col))
                if col in (6, 9, 10, 11):
                    cell.fill = profit_fill
                    cell.font = self.font_bold

        r += 1
        prof_note = ("Revenue allocation is proportional by fleet count from live ledger data. "
                     "Operating costs exclude lease principal payments. "
                     "ROI = Net Monthly Contribution ÷ Monthly Rental. "
                     "Load factor estimated from operating margin proxy "
                     f"(current net margin: {op_margin:.1%}). "
                     "AI A320neo rental converted USD→EUR at live closing rate.")
        ws.cell(row=r, column=1).value = prof_note
        ws.cell(row=r, column=1).font  = self.font_small
        ws.cell(row=r, column=1).alignment = self.align_wrap
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
        ws.row_dimensions[r].height = 24

        # ====================================================================
        # SECTION C — Lease vs Buy Analysis
        # ====================================================================
        r += 2
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        c = ws.cell(row=r, column=1)
        c.value = "C.  LEASE vs BUY ANALYSIS — NARROW-BODY (A320neo) & WIDE-BODY (B787-9)  |  IAS 16 / IFRS 16 / IAS 23"
        c.fill = self.fill_purple; c.font = self.font_white_bold; c.alignment = self.align_center

        r += 1
        c_hdrs = ["Item",
                  "A320neo — LEASE", "A320neo — BUY", "NB Difference (Lease vs Buy)",
                  "",
                  "B787-9 — LEASE", "B787-9 — BUY", "WB Difference (Lease vs Buy)",
                  "Notes / Standard"]
        for i, h in enumerate(c_hdrs, 1):
            self._hdr(ws.cell(row=r, column=i), h, fill=self.fill_purple)

        LvB_ROWS = [
            ("Aircraft List Price (USD m)",
             "N/A — lessor owns aircraft", "USD 134m (list, ~45% discount in practice)",
             "Lease: ZERO capex outlay for airline",
             "",
             "N/A — lessor owns aircraft", "USD 292m (list, ~45% discount in practice)",
             "Lease: ZERO capex outlay for airline",
             "Airbus/Boeing 2023 list prices"),
            ("Monthly Payment / Depreciation (USD)",
             "USD 355,000 (rental to lessor)", "USD 510,000* (IAS 16 straight-line dep.)",
             "Lease saves ~USD 155k/month cash",
             "",
             "USD 820,000 (rental to lessor)", "USD 1,100,000* (IAS 16 straight-line dep.)",
             "Lease saves ~USD 280k/month cash",
             "*Dep. = list price ÷ 25yr ÷ 12 (IAS 16)"),
            ("Typical Term",
             "12 years", "25 years (ownership)",
             "Lease shorter — strategic flexibility",
             "",
             "12 years", "25 years (ownership)",
             "Lease shorter — WB demand cyclicality",
             "AVAC 2023 operating lease survey"),
            ("Total Cash Cost over 12yr Term (USD m)",
             "USD 51.1m", "USD 75.0m (incl. financing at 5.25%)",
             "Lease: USD 23.9m lower over 12yr horizon",
             "",
             "USD 118.1m", "USD 162.0m (incl. financing at 5.25%)",
             "Lease: USD 43.9m lower over 12yr horizon",
             "Buy cost = price + 5.25% financing, no residual benefit"),
            ("Balance Sheet Impact (IFRS 16 / IAS 16)",
             "ROU Asset + Lease Liability on B/S", "PPE Asset + Loan Liability on B/S",
             "Both ON balance sheet post-IFRS 16 (2019+)",
             "",
             "ROU Asset + Lease Liability on B/S", "PPE Asset + Loan Liability on B/S",
             "Both ON balance sheet — old off-B/S advantage gone",
             "IFRS 16 §22; IAS 16 §7 — IFRS 16 eliminated key leasing advantage"),
            ("P&L Presentation",
             "Depreciation (opex) + Interest (finance cost)", "Depreciation (opex) + Interest (finance cost)",
             "Identical P&L structure under IFRS 16",
             "",
             "Depreciation (opex) + Interest (finance cost)", "Depreciation (opex) + Interest (finance cost)",
             "Identical",
             "IFRS 16 §49; IAS 16 §48 — front-loaded P&L impact for both"),
            ("Residual Value Risk",
             "NONE — borne entirely by lessor", "Full residual value risk to airline",
             "Lease: eliminates residual value risk",
             "",
             "NONE — borne entirely by lessor", "Full residual value risk (very high for WB)",
             "LEASE: critical advantage for wide-body",
             "IAS 16; WB residual values fell 30-50% in COVID"),
            ("Fleet Flexibility",
             "High — return aircraft at end of term", "Low — disposal costs USD 5–15m",
             "Lease: high strategic flexibility",
             "",
             "High — return aircraft at end of term", "Very Low — USD 20–40m disposal cost",
             "Lease: essential for wide-body demand cycles",
             "CAPA Fleet Analyser 2024"),
            ("Financing Cost (12yr, USD m)",
             "Embedded in rental (~4.5% lessor cost of capital)", "5.25% p.a. on USD 134m = USD 42m",
             "Buy: explicit financing cost; Lease: implicit via lessor margin",
             "",
             "Embedded in rental (~4.8% lessor cost of capital)", "5.25% p.a. on USD 292m = USD 92m",
             "Sub-IG carrier: buy financing significantly more expensive",
             "SOFR/EURIBOR + credit spread; BBB airline profile"),
            ("Maintenance Reserves",
             "USD 50–100k/month to lessor (additional to rent)", "Self-funded; provisioned per IAS 37",
             "Lease: cash reserve to lessor; Buy: IAS 37 provision",
             "",
             "USD 150–300k/month to lessor", "Self-funded per IAS 37",
             "Wide-body maintenance significantly more complex & costly",
             "IAS 37 §14; AVAC MR rates 2024"),
            ("Tax Treatment (IFRS jurisdiction)",
             "Deductible: ROU depreciation + interest expense", "Deductible: PPE depreciation + interest expense",
             "Near-identical tax treatment post-IFRS 16",
             "",
             "Deductible: ROU depreciation + interest expense", "Deductible: PPE depreciation + interest expense",
             "Near-identical",
             "IAS 12 §41A — initial recognition exemption for ROU/lease liability"),
        ]

        for i, row_data in enumerate(LvB_ROWS):
            r += 1
            self._alt_row(ws, r, 9)
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=col)
                cell.value = val
                cell.font  = self.font_bold if col == 1 else self.font_normal
                cell.border = self.border_thin
                cell.alignment = self.align_wrap
            ws.row_dimensions[r].height = 36

        # Recommendation rows
        rec_rows = [
            ("RECOMMENDATION — A320neo (Narrow-Body)",
             "LEASE preferred for sub-investment-grade carriers (IBR > 5%); BUY considered only by SIA/Lufthansa with very low cost of capital",
             "", "Key driver: financing cost + zero capex",
             "", "N/A", "N/A", "N/A",
             "Rating, IBR, network flexibility, fleet age are decisive"),
            ("RECOMMENDATION — B787-9 (Wide-Body)",
             "LEASE strongly preferred across ALL airline credit profiles",
             "", "Residual value risk on WB is extreme; demand cycles volatile",
             "", "N/A", "N/A", "N/A",
             "Even IG airlines (Emirates, Cathay) prefer to lease wide-bodies"),
        ]
        for rec in rec_rows:
            r += 1
            for col, val in enumerate(rec, 1):
                cell = ws.cell(row=r, column=col)
                cell.value = val
                cell.fill  = self.fill_gold if col == 1 else self.fill_amber
                cell.font  = Font(name="Calibri", bold=True, size=10,
                                  color="000000" if col == 1 else "000000")
                cell.border = self.border_thin
                cell.alignment = self.align_wrap
            ws.row_dimensions[r].height = 30

        # ====================================================================
        # SECTION D — Aircraft Lease Market Rates
        # ====================================================================
        r += 2
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        c = ws.cell(row=r, column=1)
        c.value = "D.  AIRCRAFT LEASE MARKET RATES — INDUSTRY BENCHMARKS  (AVAC / IBA / CAPA, Q1 2024)"
        c.fill = self.fill_teal; c.font = self.font_white_bold; c.alignment = self.align_center

        r += 1
        d_hdrs = ["Aircraft Type", "Category", "Monthly Lease Rate (USD)",
                  "Typical Term (yrs)", "Major Lessor(s)",
                  "Current Availability", "Pre-COVID Rate (USD)", "Change vs Pre-COVID %"]
        for i, h in enumerate(d_hdrs, 1):
            self._hdr(ws.cell(row=r, column=i), h, fill=self.fill_teal)

        MARKET_RATES = [
            # (type, category, monthly_usd, term_yr, lessors, availability, pre_covid)
            ("A320neo",    "Narrow-Body", 355_000, 12, "AerCap / SMBC / GECAS",    "Tight",      290_000),
            ("A321neo",    "Narrow-Body", 430_000, 12, "AerCap / Air Lease Corp",   "Very Tight", 340_000),
            ("B737 MAX 8", "Narrow-Body", 340_000, 12, "SMBC / GECAS / BOC Avn.",  "Tight",      270_000),
            ("B737 MAX 10","Narrow-Body", 395_000, 12, "SMBC / BOC Aviation",       "Moderate",   None),
            ("A350-900",   "Wide-Body",   850_000, 12, "AerCap / Air Lease Corp",   "Moderate",   780_000),
            ("B787-9",     "Wide-Body",   820_000, 12, "AerCap / GECAS / Avolon",  "Moderate",   750_000),
            ("A380",       "Wide-Body",   850_000, 10, "AerCap (limited market)",   "Loose",      1_200_000),
        ]

        avail_fill_map = {
            "Very Tight": self.fill_red,
            "Tight":      self.fill_red,
            "Moderate":   self.fill_amber,
            "Loose":      self.fill_green_l,
        }

        for i, (atype, cat, monthly, term, lessors, avail, pre_covid) in enumerate(MARKET_RATES):
            r += 1
            self._alt_row(ws, r, 8)
            for col, val in enumerate([atype, cat, monthly, term, lessors, avail, "", ""], 1):
                cell = ws.cell(row=r, column=col)
                if col == 3:
                    self._data(cell, val, "#,##0")
                elif col == 1:
                    cell.value = val; cell.font = self.font_bold; cell.border = self.border_thin
                else:
                    self._data(cell, val)
                if col == 6:
                    cell.fill = avail_fill_map.get(avail, self.fill_alt1)
                    cell.font = self.font_bold

            # Pre-COVID rate (col 7)
            cell7 = ws.cell(row=r, column=7)
            if pre_covid is not None:
                self._data(cell7, pre_covid, "#,##0")
            else:
                cell7.value = "N/A (new type)"
                cell7.font  = self.font_small
                cell7.border = self.border_thin

            # Change vs pre-COVID (col 8)
            cell8 = ws.cell(row=r, column=8)
            if pre_covid is not None:
                pct = (monthly - pre_covid) / pre_covid
                self._data(cell8, pct, "0.0%")
                cell8.fill = self.fill_red if pct > 0 else self.fill_green_l
                cell8.font = self.font_bold
            else:
                cell8.value = "—"
                cell8.font  = self.font_small
                cell8.border = self.border_thin

        r += 1
        mkt_note = ("Sources: AVAC Aircraft Values & Lease Rates Q1 2024; IBA Aviation Insight Mar 2024; "
                    "CAPA Fleet Analyser. Rates are mid-market monthly rental for 12-year operating lease, "
                    "new-delivery aircraft, BBB-rated lessee. Actual rates vary ±15–20% by lessee credit, "
                    "delivery slot, and term. Pre-COVID = Q4 2019. Change % = post-COVID rate uplift driven "
                    "by delivery backlogs (Airbus/Boeing) and surge in narrow-body demand post-2022.")
        ws.cell(row=r, column=1).value = mkt_note
        ws.cell(row=r, column=1).font  = self.font_small
        ws.cell(row=r, column=1).alignment = self.align_wrap
        ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=8)
        ws.row_dimensions[r].height = 28

        self._last_updated(ws)

        # Manual column widths
        col_widths = [34, 18, 22, 20, 12, 22, 14, 16, 24, 48, 18, 18]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
class RealtimeLoop:
    def __init__(self, interval: int, output_file: str, use_live: bool):
        self.interval    = interval
        self.output_file = output_file
        self.fx          = FXRateManager(use_live=use_live)
        self.ledger      = LedgerBook()
        self.leases      = self._init_leases()
        self.writer      = WorkbookWriter(self.ledger, self.fx, self.leases)
        self._cycle      = 0
        self._last_date  = datetime.date(_today.year, 1, 15)  # seed date for new txns

    def _init_leases(self) -> List[IFRS16LeaseSchedule]:
        return [
            IFRS16LeaseSchedule(LeaseConfig(
                lease_id="LEASE-A320-01", aircraft_type="Airbus A320neo",
                commencement=datetime.date(_today.year - 2, 4, 1),
                lease_term_months=120, monthly_payment=Decimal("210000"),
                effective_rate=Decimal("0.065"))),
            IFRS16LeaseSchedule(LeaseConfig(
                lease_id="LEASE-B737-01", aircraft_type="Boeing 737 MAX 8",
                commencement=datetime.date(_today.year - 1, 7, 1),
                lease_term_months=96, monthly_payment=Decimal("185000"),
                effective_rate=Decimal("0.060"))),
        ]

    def _make_tx(self, desc, account, atype, dr_eur, cr_eur, ccy, rate,
                 monetary=True, lease_id=None, date_offset_days=0) -> Transaction:
        d = self._last_date + datetime.timedelta(days=date_offset_days)
        return Transaction(
            tx_id=self.ledger.next_tx_id(),
            date=d,
            description=desc,
            account=account,
            account_type=atype,
            debit_eur=Decimal(str(dr_eur)),
            credit_eur=Decimal(str(cr_eur)),
            currency=ccy,
            fx_rate=Decimal(str(rate)),
            is_monetary=monetary,
            lease_id=lease_id,
        )

    def _preload_transactions(self):
        rates = _BASE_RATES
        # Each entry: (desc, account, atype, dr, cr, ccy, rate, is_monetary, day_offset, lease_id)
        txs_def = [
            # Fuel purchase cycle
            ("Jet fuel purchase — JFK refuelling",   "Fuel Expense",         "EXPENSE",   480000, 0,       "USD", rates["USD"], True,  0,  None),
            ("Fuel supplier AP accrual",             "Fuel Payable",         "LIABILITY",  0, 480000,      "USD", rates["USD"], True,  0,  None),
            ("Fuel AP payment cleared",              "Cash/Bank",            "ASSET",      0, 480000,      "USD", rates["USD"], True,  1,  None),
            ("Fuel AP settled",                      "Fuel Payable",         "LIABILITY",  480000, 0,      "USD", rates["USD"], True,  1,  None),
            # Lease A320
            ("Aircraft lease payment — A320neo Q1",  "Lease Liability",      "LIABILITY",  210000, 0,      "EUR", 1.0,          True,  2,  "LEASE-A320-01"),
            ("Lease interest charge — A320neo Q1",   "Interest Expense",     "EXPENSE",    14300, 0,       "EUR", 1.0,          False, 2,  "LEASE-A320-01"),
            ("ROU asset depreciation — A320neo",     "ROU Asset Depreciation","EXPENSE",   38500, 0,       "EUR", 1.0,          False, 2,  None),
            # Maintenance
            ("Heavy maintenance reserve — Heathrow", "Maintenance Reserve",  "EXPENSE",    95000, 0,       "GBP", rates["GBP"], False, 3,  None),
            ("Maintenance payable accrual",          "Maintenance Payable",  "LIABILITY",  0, 95000,       "GBP", rates["GBP"], True,  3,  None),
            # Revenue
            ("Passenger revenue — LHR-SIN route",    "Passenger Revenue",    "REVENUE",    0, 620000,      "SGD", rates["SGD"], False, 4,  None),
            ("Cargo revenue — BOM-FRA freight",      "Cargo Revenue",        "REVENUE",    0, 8500000,     "INR", rates["INR"], False, 4,  None),
            ("Passenger revenue — MUC-DXB",          "Passenger Revenue",    "REVENUE",    0, 380000,      "USD", rates["USD"], False, 5,  None),
            # Staff
            ("Staff salaries — flight crew",         "Staff Costs",          "EXPENSE",    340000, 0,      "EUR", 1.0,          False, 6,  None),
            # Airport / ATC
            ("Airport landing charges — CDG",        "Airport Charges",      "EXPENSE",    12400, 0,       "EUR", 1.0,          False, 7,  None),
            ("Airport charges — LHR",                "Airport Charges",      "EXPENSE",    18600, 0,       "GBP", rates["GBP"], False, 7,  None),
            ("ATC charges payable — Eurocontrol",    "ATC Charges Payable",  "LIABILITY",  0, 7200,        "EUR", 1.0,          True,  7,  None),
            # Ground handling
            ("Ground handling — Singapore Changi",   "Ground Handling",      "EXPENSE",    42000, 0,       "SGD", rates["SGD"], False, 8,  None),
            # Other opex
            ("Insurance premium — fleet",            "Insurance Expense",    "EXPENSE",    56000, 0,       "EUR", 1.0,          False, 9,  None),
            ("Catering costs — economy class",       "Catering Costs",       "EXPENSE",    28400, 0,       "EUR", 1.0,          False, 9,  None),
            ("Fuel hedge gain settlement",           "Fuel Hedge Gain",      "REVENUE",    0, 22000,       "USD", rates["USD"], False, 10, None),
            ("Navigation fees — Eurocontrol Nov",    "Navigation Fees",      "EXPENSE",    9800, 0,        "EUR", 1.0,          False, 11, None),
            # Lease B737
            ("B737 lease payment Q1",                "Lease Liability",      "LIABILITY",  185000, 0,      "EUR", 1.0,          True,  12, "LEASE-B737-01"),
            ("B737 lease interest Q1",               "Interest Expense",     "EXPENSE",    10800, 0,       "EUR", 1.0,          False, 12, "LEASE-B737-01"),
        ]
        for desc, account, atype, dr, cr, ccy, rate, is_mon, day_offset, lease_id in txs_def:
            tx = self._make_tx(desc, account, atype, dr, cr, ccy, float(rate), is_mon, lease_id, day_offset)
            self.ledger.add_transaction(tx)
        self._last_date += datetime.timedelta(days=14)

    def _generate_random_transaction(self) -> Transaction:
        closing = self.fx.closing_rates()
        self._last_date += datetime.timedelta(days=random.randint(0, 3))
        templates = [
            ("Supplemental fuel uplift — route",  "Fuel Expense",   "EXPENSE",  "USD", closing["USD"], 120000, True),
            ("Passenger revenue — charter",        "Charter Revenue","REVENUE",  "GBP", closing["GBP"], 0,      False),
            ("Cargo uplift — FRA-BOM",             "Cargo Revenue",  "REVENUE",  "INR", closing["INR"], 0,      False),
            ("Ground handling — Heathrow",         "Ground Handling","EXPENSE",  "GBP", closing["GBP"], 50000,  False),
            ("Ancillary revenue — seat upgrades",  "Ancillary Revenue","REVENUE","SGD", closing["SGD"], 0,      False),
            ("Catering replenishment",             "Catering Costs", "EXPENSE",  "EUR", Decimal("1"),   18000,  False),
            ("Airport slot fee — LHR",             "Airport Charges","EXPENSE",  "GBP", closing["GBP"], 30000,  False),
            ("Navigation overflight charges",      "Navigation Fees","EXPENSE",  "EUR", Decimal("1"),   5500,   False),
        ]
        tmpl = random.choice(templates)
        desc, acct, atype, ccy, rate, base_eur, is_mon = tmpl
        factor = Decimal(str(1 + random.uniform(-0.15, 0.15)))
        if base_eur > 0:
            amt_eur = _d2(Decimal(str(base_eur)) * factor)
        else:
            amt_eur = _d2(Decimal(str(random.randint(50000, 500000))) * factor)

        if atype == "EXPENSE":
            dr, cr = float(amt_eur), 0.0
        else:
            dr, cr = 0.0, float(amt_eur)

        return self._make_tx(
            f"{desc} [{self._last_date.strftime('%b %d')}]",
            acct, atype, dr, cr, ccy, float(rate), is_mon
        )

    def run(self):
        print("=" * 70)
        print("  AIRLINE ACCOUNTING REAL-TIME EXCEL GENERATOR")
        print(f"  Standards: IFRS 16 | IAS 21/IFRS 21 | IATA AAG | Ind AS 116/11")
        print(f"  Base currency: EUR  |  Parallel ledgers: USD, GBP, INR, SGD")
        print(f"  Output: {self.output_file}  |  Interval: {self.interval}s")
        print("  Press Ctrl+C to stop.")
        print("=" * 70)

        self._preload_transactions()

        while True:
            self._cycle += 1
            try:
                rates = self.fx.fetch_or_simulate()
                self.fx.record_rates(rates)

                n_new = random.randint(1, 2)
                for _ in range(n_new):
                    tx = self._generate_random_transaction()
                    self.ledger.add_transaction(tx)

                self.writer.write(self.output_file)
                ts  = datetime.datetime.now().strftime("%H:%M:%S")
                txn = len(self.ledger.all_transactions())
                ni  = float(self.ledger.net_income())
                print(f"[{ts}] Cycle {self._cycle:4d} | Txns: {txn:4d} | "
                      f"Net Income: EUR {ni:>14,.0f} | "
                      f"USD {float(rates['USD']):.4f} | GBP {float(rates['GBP']):.4f} | "
                      f"INR {float(rates['INR']):.2f} | SGD {float(rates['SGD']):.4f} | "
                      f"{'LIVE' if self.fx.is_live() else 'SIM ':4s}")

            except PermissionError:
                print(f"[WARN] {self.output_file} is open in Excel — skipping write (data accumulating in memory)")
            except KeyboardInterrupt:
                raise
            except Exception as exc:
                print(f"[ERROR] Cycle {self._cycle}: {exc}")

            time.sleep(self.interval)


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Airline Accounting Real-Time Excel Generator\n"
                    "Standards: IFRS 16, IAS 21/IFRS 21, IATA AAG, Ind AS 116/11",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--interval", type=int, default=REFRESH_INTERVAL,
                        help=f"Refresh interval in seconds (default: {REFRESH_INTERVAL})")
    parser.add_argument("--output",   type=str, default=OUTPUT_FILE,
                        help=f"Output Excel file path (default: {OUTPUT_FILE})")
    parser.add_argument("--no-api",   action="store_true",
                        help="Skip live FX API, use simulation only")
    args = parser.parse_args()

    loop = RealtimeLoop(
        interval=args.interval,
        output_file=args.output,
        use_live=not args.no_api,
    )
    try:
        loop.run()
    except KeyboardInterrupt:
        print("\n[INFO] Stopped by user.")
        sys.exit(0)


if __name__ == "__main__":
    main()
